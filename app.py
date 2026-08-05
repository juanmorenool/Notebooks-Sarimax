import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from scipy import stats
import json
import os
import base64
import openpyxl
from pathlib import Path
from datetime import datetime
from io import BytesIO
import tempfile
import matplotlib
matplotlib.use("Agg")  # Backend sin interfaz grafica - obligatorio en servidores (Streamlit Cloud)
import matplotlib.pyplot as plt
from fpdf import FPDF
from openpyxl.drawing.image import Image as XLImage

# =============================================================================
# PALETA CORPORATIVA BANCA
# =============================================================================
NAVY    = "#0B2545"
BLUE    = "#1E5AA8"
GREEN   = "#1A6B3E"
RED     = "#B83232"
GRAY    = "#6C757D"
LTGRAY  = "#ADB5BD"
BG      = "#F8F9FA"
WHITE   = "#FFFFFF"
BORDER  = "#DEE2E6"
TEXT    = "#1A1D23"
MUTED   = "#5A6270"
TINT    = "#EDF2F7"

ESTADO_OK      = ("#E8F5E9", "#1A6B3E", "CUMPLE")
ESTADO_WARN    = ("#FFF8E1", "#B8860B", "REVISAR")
ESTADO_FAIL    = ("#FFEBEE", "#B83232", "NO CUMPLE")
ESTADO_NEUTRAL = ("#F5F5F5", "#5A6270", "N/A")

SCORE_COLORS = {
    'A': ("#E8F5E9", "#1A6B3E"),
    'B': ("#E3F2FD", "#1565C0"),
    'C': ("#FFF8E1", "#B8860B"),
    'D': ("#FFEBEE", "#B83232"),
}

SCORE_NUM_MAP = {'A': 10.0, 'B': 7.5, 'C': 5.0, 'D': 2.5}
PESOS_SCORE_GLOBAL = {'ljung_box': 0.40, 'jarque_bera': 0.30, 'heterocedasticidad': 0.30}

PAISES_MAP = {
    'colombia': 'Colombia',
    'panama': 'Panama',
    'panamá': 'Panama',
    'costa rica': 'Costa Rica',
    'co': 'Colombia',
    'pa': 'Panama',
    'cr': 'Costa Rica',
}

BANDERAS_PAISES = {
    'Colombia': 'co',
    'Panama': 'pa',
    'Costa Rica': 'cr',
    'co': 'co', 'pa': 'pa', 'cr': 'cr',
    'CO': 'co', 'PA': 'pa', 'CR': 'cr',
}

CARTERAS_MAP = {
    'vivi': 'Vivienda', 'vivienda': 'Vivienda',
    'cons': 'Consumo', 'consumo': 'Consumo',
    'com': 'Comercial', 'comercial': 'Comercial',
    'micro': 'Microcredito',
}

# --- Mapeos para el generador ---
PAIS_MAP_GEN = {
    'Colombia': 'CO',
    'Panama': 'PA',
    'Costa Rica': 'CR',
}

CARTERA_MAP_GEN = {
    'Vivienda': 'vivienda',
    'Consumo': 'consumo',
    'Comercial': 'comercial',
    'Microcredito': 'microcredito',
    'Corporativo': 'corporativo',
    'Pyme': 'pyme',
    'Tarjeta': 'tarjeta',
    'Vehiculo': 'vehiculo',
}

CARTERA_LABEL_GEN = {
    'vivienda': 'Vivienda',
    'consumo': 'Consumo',
    'comercial': 'Comercial',
    'microcredito': 'Microcredito',
    'corporativo': 'Corporativo',
    'pyme': 'Pyme',
    'tarjeta': 'Tarjeta',
    'vehiculo': 'Vehiculo',
}

# =============================================================================
# TEXTOS DOCUMENTO METODOLOGICO
# =============================================================================
TEXTO_METODOLOGIA_SARIMAX = (
    "Este documento resume la seleccion final de un modelo SARIMAX para IFRS 9. "
    "El motor sigue 8 bloques: (1) carga y validacion de insumos, (2) limpieza y "
    "transformaciones, (3) construccion de escenarios, (4) busqueda de rezagos AR/MA, "
    "(5) ajuste econometrico, (6) diagnosticos de residuos, (7) filtros tecnicos y "
    "(8) exportacion con trazabilidad."
)

GLOSARIO_FWL = "FWL [CAMPO]: factor de ajuste por escenario macro utilizado para sensibilidad y consistencia del modelo."
GLOSARIO_LOGIT = "LOGIT [CAMPO]: transformacion logit de la endogena para estabilizar escala y mejorar ajuste."
GLOSARIO_MODO = "MODO [CAMPO]: estrategia de construccion de la endogena (actual o media movil)."
GLOSARIO_SENSIBILIDAD = "SENSIBILIDAD [CAMPO]: diferencia entre medias de escenarios OPT y ADV para validar reaccion del modelo."
GLOSARIO_PARAMETROS = "PARAMETROS [CAMPO]: conjunto de hiperparametros del motor (max lags, VIF, top exportar y rangos FWL)."

def inject_css():
    st.markdown(f"""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    html, body, [class*="css"], .stApp, .stMarkdown, .stDataFrame, .stButton,
    .stSelectbox, .stTextInput, .stNumberInput, .stTabs {{
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
    }}
    #MainMenu {{ visibility: hidden; }}
    footer {{ visibility: hidden; }}
    header[data-testid="stHeader"] {{ background: transparent; }}
    .block-container {{ padding-top: 0.5rem; padding-bottom: 2rem; max-width: 1400px; }}
    .stApp {{ background-color: {BG}; }}
    section[data-testid="stSidebar"] {{
        background-color: {WHITE}; border-right: 1px solid {BORDER};
        min-width: 260px !important; max-width: 260px !important;
    }}
    /* Compactar file uploader dentro del sidebar */
    section[data-testid="stSidebar"] [data-testid="stFileUploader"] {{
        padding: 0 !important;
    }}
    section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] {{
        min-height: 50px !important;
        padding: 8px 6px !important;
    }}
    section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] div div small {{
        font-size: 11px !important;
        line-height: 1.2 !important;
    }}
    section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button {{
        font-size: 11px !important;
        padding: 2px 8px !important;
    }}
    section[data-testid="stSidebar"] .block-container {{ padding: 16px; }}
    h1 {{ color: {NAVY} !important; font-weight: 700 !important; font-size: 22px !important; }}
    h2 {{ color: {NAVY} !important; font-weight: 600 !important; font-size: 16px !important; margin-top: 0.2rem !important; }}
    h3 {{ color: {NAVY} !important; font-weight: 600 !important; font-size: 14px !important;
         border-left: 3px solid {BLUE}; padding-left: 10px; margin-top: 0.3rem !important; }}
    .stTabs [data-baseweb="tab-list"] {{ gap: 2px; border-bottom: 1px solid {BORDER}; }}
    .stTabs [data-baseweb="tab"] {{
        background-color: transparent; border-radius: 6px 6px 0 0;
        padding: 8px 16px; color: {MUTED}; font-weight: 500; font-size: 13px;
    }}
    .stTabs [aria-selected="true"] {{
        background-color: {TINT} !important; color: {NAVY} !important; font-weight: 600;
    }}
    div[data-testid="stDataFrame"] {{ border: 1px solid {BORDER}; border-radius: 6px; }}
    section[data-testid="stSidebar"] {{ position: sticky; top: 0; height: 100vh; overflow-y: auto; }}
    .st-key-kb_hidden {{
        position: fixed !important; top: -9999px !important; left: -9999px !important;
        height: 1px !important; width: 1px !important; overflow: hidden !important;
    }}
    section[data-testid="stSidebar"] button[kind="secondary"][class*="st-key-btn_sec_"],
    section[data-testid="stSidebar"] .st-key-btn_sec_contexto button,
    section[data-testid="stSidebar"] .st-key-btn_sec_orden button,
    section[data-testid="stSidebar"] .st-key-btn_sec_exogenas button {{
        background: none !important; border: none !important; box-shadow: none !important;
        padding: 4px 0 !important; text-align: left !important; justify-content: flex-start !important;
        font-size: 13px !important; font-weight: 700 !important; color: {NAVY} !important;
        text-transform: uppercase; letter-spacing: 0.5px;
    }}
    section[data-testid="stSidebar"] .st-key-btn_sec_contexto button:hover,
    section[data-testid="stSidebar"] .st-key-btn_sec_orden button:hover,
    section[data-testid="stSidebar"] .st-key-btn_sec_exogenas button:hover {{
        color: {BLUE} !important;
    }}
    /* Estilo para botones primarios del generador */
    button[kind="primary"] {{
        background-color: {BLUE} !important;
        color: {WHITE} !important;
        border-radius: 6px !important;
        font-weight: 600 !important;
    }}
    button[kind="primary"]:hover {{
        background-color: {NAVY} !important;
    }}
    </style>
    """, unsafe_allow_html=True)

def pill(text, color_bg, color_fg):
    return f'<span style="background:{color_bg};color:{color_fg};font-size:11px;padding:2px 8px;border-radius:4px;font-weight:600;text-transform:uppercase;letter-spacing:0.4px;">{text}</span>'

def tag_ok():    return pill("CUMPLE", ESTADO_OK[0], ESTADO_OK[1])
def tag_warn():  return pill("REVISAR", ESTADO_WARN[0], ESTADO_WARN[1])
def tag_fail():  return pill("NO CUMPLE", ESTADO_FAIL[0], ESTADO_FAIL[1])
def tag_neutral(): return pill("N/A", ESTADO_NEUTRAL[0], ESTADO_NEUTRAL[1])

def card_kpi(title, value, subtitle="", accent=NAVY):
    sub = f'<p style="font-size:12px;color:{MUTED};margin:4px 0 0;line-height:1.3;">{subtitle}</p>' if subtitle else ''
    return f"""
    <div style="background:{WHITE};border:1px solid {BORDER};border-radius:8px;padding:14px 16px;height:100%;box-sizing:border-box;">
        <p style="font-size:10px;color:{LTGRAY};margin:0 0 6px;text-transform:uppercase;letter-spacing:0.6px;font-weight:600;">{title}</p>
        <p style="font-size:18px;font-weight:700;color:{accent};margin:0;line-height:1.2;">{value}</p>
        {sub}
    </div>
    """

def card_metric(label, value, color=TEXT):
    return f"""
    <div style="background:{WHITE};border:1px solid {BORDER};border-radius:6px;padding:12px 14px;">
        <p style="font-size:10px;color:{LTGRAY};margin:0 0 4px;text-transform:uppercase;letter-spacing:0.6px;font-weight:600;">{label}</p>
        <p style="font-size:20px;font-weight:700;color:{color};margin:0;">{value}</p>
    </div>
    """

def divider():
    return f"<div style='height:1px;background:{BORDER};margin:16px 0;'></div>"

def section_title(text):
    return f"<p style='font-size:13px;font-weight:700;color:{NAVY};margin:0 0 12px;text-transform:uppercase;letter-spacing:0.5px;'>{text}</p>"

def obtener_bandera_pais(pais, ancho=20):
    if not pais:
        return ""
    pais_str = str(pais).strip()
    codigo = BANDERAS_PAISES.get(pais_str)
    if not codigo:
        codigo = BANDERAS_PAISES.get(pais_str.lower())
    if not codigo:
        return ""
    return (f'<img src="https://flagcdn.com/w40/{codigo}.png" width="{ancho}" '
            f'style="vertical-align:middle;border-radius:2px;margin-right:6px;box-shadow:0 0 0 1px {BORDER};" '
            f'alt="{pais_str}">')

# =============================================================================
# PREFERENCIAS DEL SIDEBAR (persistentes entre sesiones)
# =============================================================================
PREFS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".sidebar_prefs.json")

CLAVES_PREFS_SIDEBAR = [
    "sec_contexto", "sec_orden", "sec_exogenas",
    "criterio_ordenamiento",
    "filtro_ljung", "filtro_jarque", "filtro_hetero",
    "filtro_favoritos", "nav_sticky",
]

def cargar_prefs_sidebar():
    try:
        with open(PREFS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def guardar_prefs_sidebar():
    prefs = {clave: st.session_state.get(clave) for clave in CLAVES_PREFS_SIDEBAR}
    try:
        with open(PREFS_PATH, "w", encoding="utf-8") as f:
            json.dump(prefs, f)
    except Exception:
        pass

def encabezado_colapsable(titulo, key):
    if key not in st.session_state:
        st.session_state[key] = True
    abierto = st.session_state[key]
    icono = "-" if abierto else "+"
    if st.button(f"{icono}  {titulo}", key=f"btn_{key}", use_container_width=True):
        st.session_state[key] = not abierto
        guardar_prefs_sidebar()
        st.rerun()
    return st.session_state[key]

# =============================================================================
# PARSER
# =============================================================================
def convertir_fecha(serie):
    if serie is None or len(serie) == 0:
        return serie
    if pd.api.types.is_datetime64_any_dtype(serie):
        return serie
    try:
        numeric = pd.to_numeric(serie, errors='coerce')
        if numeric.notna().sum() > 0 and numeric.dropna().min() > 30000:
            result = pd.to_datetime(numeric, unit='D', origin='1899-12-30', errors='coerce')
            if result.notna().sum() > 0:
                return result
    except:
        pass
    return pd.to_datetime(serie, errors='coerce')

def leer_meta_embebida(file, prefix="sarimax_meta"):
    try:
        file.seek(0)
        wb = openpyxl.load_workbook(file, read_only=True)
        props = wb.custom_doc_props
        n_prop_name = f"{prefix}_n"
        if n_prop_name not in props.names:
            return None
        n_partes = int(props[n_prop_name].value)
        partes = [props[f"{prefix}_{idx:02d}"].value for idx in range(1, n_partes + 1)]
        return json.loads("".join(partes))
    except Exception:
        return None
    finally:
        file.seek(0)

def parsear_excel(file):
    xls = pd.ExcelFile(file)
    modelos = {}
    for sheet_name in xls.sheet_names:
        try:
            df_raw = pd.read_excel(xls, sheet_name=sheet_name, header=None)
        except Exception:
            continue
        if len(df_raw) < 2:
            continue
        headers = [str(v).strip() if pd.notna(v) else "" for v in df_raw.iloc[1].values]
        col_map = {}
        for idx, name in enumerate(headers):
            if name:
                col_map.setdefault(name, []).append(idx)
        modelo = {"nombre": sheet_name}
        exog_cols, exog_names = [], set()
        for idx, name in enumerate(headers):
            nu = name.upper()
            if nu in ['FECHA', 'BASE', 'ADVERSO', 'OPTIMISTA'] or nu.startswith('FWL'):
                continue
            if nu.endswith(('_BASE', '_ADVERSO', '_OPTIMISTA')):
                exog_cols.append(name)
                base_name = name
                for suffix in ['_BASE', '_ADVERSO', '_OPTIMISTA']:
                    if base_name.upper().endswith(suffix):
                        base_name = base_name[:-len(suffix)]
                        break
                exog_names.add(base_name)
        endogena_cols = ['BASE', 'ADVERSO', 'OPTIMISTA']
        cols_seccion1 = ['fecha'] + endogena_cols + exog_cols
        if 'FWL_BASE' in col_map:
            cols_seccion1.extend(['FWL_BASE', 'FWL_ADVERSO', 'FWL_OPTIMISTA'])
        cols_seccion1 = [c for c in cols_seccion1 if c in col_map]
        idx_seccion1 = [col_map[c][0] for c in cols_seccion1]
        df_seccion1 = df_raw.iloc[2:, idx_seccion1].copy()
        df_seccion1.columns = cols_seccion1
        df_seccion1 = df_seccion1.dropna(how='all').reset_index(drop=True)
        if 'fecha' in df_seccion1.columns:
            df_seccion1['fecha'] = convertir_fecha(df_seccion1['fecha'])
        modelo['fecha_endogena'] = df_seccion1
        modelo['endogenas_cols'] = endogena_cols
        modelo['exogenas_cols'] = exog_cols
        modelo['exogenas_nombres'] = sorted(list(exog_names))
        if exog_cols:
            modelo['exogenas'] = df_seccion1[['fecha'] + exog_cols] if 'fecha' in df_seccion1.columns else df_seccion1[exog_cols]
        else:
            modelo['exogenas'] = None
        fwl_cols = [c for c in ['FWL_BASE', 'FWL_ADVERSO', 'FWL_OPTIMISTA'] if c in col_map]
        if fwl_cols and 'fecha' in col_map:
            idx_fwl = [col_map['fecha'][0]] + [col_map[c][0] for c in fwl_cols]
            df_fwl = df_raw.iloc[2:, idx_fwl].copy()
            df_fwl.columns = ['fecha'] + fwl_cols
            df_fwl = df_fwl.dropna(how='all').reset_index(drop=True)
            df_fwl['fecha'] = convertir_fecha(df_fwl['fecha'])
            df_fwl = df_fwl.dropna(subset=['FWL_BASE']).reset_index(drop=True)
            modelo['fwl_12m'] = df_fwl
        else:
            modelo['fwl_12m'] = None
        if 'Ano' in col_map and 'Escenario' in col_map and 'Factor FWL' in col_map:
            idx_anual = [col_map['Ano'][0], col_map['Escenario'][0], col_map['Factor FWL'][0]]
            df_anual = df_raw.iloc[2:, idx_anual].copy()
            df_anual.columns = ['Ano', 'Escenario', 'Factor FWL']
            df_anual = df_anual.dropna(how='all').reset_index(drop=True)
            modelo['fwl_anual'] = df_anual
        else:
            modelo['fwl_anual'] = None
        if 'Obs' in col_map and 'Residuo' in col_map:
            idx_res = [col_map['Obs'][0], col_map['Residuo'][0]]
            df_res = df_raw.iloc[2:, idx_res].copy()
            df_res.columns = ['Obs', 'Residuo']
            df_res = df_res.dropna(how='all').reset_index(drop=True)
            modelo['residuos_ind'] = df_res
        else:
            modelo['residuos_ind'] = None
        if 'Estadistico' in col_map and 'Valor' in col_map:
            idx_est = col_map['Estadistico'][0]
            idx_val = col_map['Valor'][0]
            df_resumen = df_raw.iloc[2:, [idx_est, idx_val]].copy()
            df_resumen.columns = ['Estadistico', 'Valor']
            df_resumen = df_resumen.dropna(how='all').reset_index(drop=True)
            modelo['resumen_residuos'] = df_resumen
        else:
            modelo['resumen_residuos'] = None
        if 'Variable' in col_map and 'Coeficiente' in col_map and 'P_value' in col_map:
            idx_var = col_map['Variable'][0]
            idx_coef = col_map['Coeficiente'][0]
            idx_pval = col_map['P_value'][0]
            df_coef = df_raw.iloc[2:, [idx_var, idx_coef, idx_pval]].copy()
            df_coef.columns = ['Variable', 'Coeficiente', 'P_value']
            df_coef = df_coef.dropna(how='all').reset_index(drop=True)
            modelo['coeficientes'] = df_coef
        else:
            modelo['coeficientes'] = None
        if 'Prueba' in col_map and 'Estadistico' in col_map and 'P_value' in col_map:
            idx_prueba = col_map['Prueba'][0]
            idx_est = col_map['Estadistico'][-1]
            idx_pval = col_map['P_value'][-1]
            df_pruebas = df_raw.iloc[2:, [idx_prueba, idx_est, idx_pval]].copy()
            df_pruebas.columns = ['Prueba', 'Estadistico', 'P_value']
            df_pruebas = df_pruebas.dropna(how='all').reset_index(drop=True)
            modelo['pruebas'] = df_pruebas
        else:
            modelo['pruebas'] = None
        modelo['observaciones'] = len(modelo['fecha_endogena'].dropna(how='all')) if modelo['fecha_endogena'] is not None and not modelo['fecha_endogena'].empty else 0
        modelos[sheet_name] = modelo
    return modelos

# =============================================================================
# UTILIDADES
# =============================================================================
def contar_pruebas_aprobadas(pruebas_df):
    if pruebas_df is None or pruebas_df.empty:
        return 0, 3
    aprobadas = 0
    for _, row in pruebas_df.iterrows():
        prueba = str(row.get('Prueba', '')).lower()
        p_val = row.get('P_value', None)
        if p_val is None or pd.isna(p_val):
            continue
        try: p_val = float(p_val)
        except: continue
        if 'ljung' in prueba or 'box' in prueba:
            if p_val > 0.05: aprobadas += 1
        elif 'jarque' in prueba or 'bera' in prueba:
            if p_val > 0.05: aprobadas += 1
        elif 'hetero' in prueba or 'arch' in prueba:
            if p_val > 0.05: aprobadas += 1
    return aprobadas, 3

def clasificar_variable(var_name):
    var_lower = str(var_name).lower()
    if var_lower.startswith('ar.'): return 'AR'
    elif var_lower.startswith('ma.'): return 'MA'
    elif var_lower == 'intercept': return 'Exogena'
    elif var_lower.startswith('var_'): return 'Exogena'
    elif var_lower == 'sigma2': return 'Varianza'
    return 'Otro'

def contar_ar_ma(coeficientes_df):
    if coeficientes_df is None or coeficientes_df.empty:
        return 0, 0
    ar_count = 0
    ma_count = 0
    for _, row in coeficientes_df.iterrows():
        var = str(row.get('Variable', '')).lower()
        if var.startswith('ar.'):
            ar_count += 1
        elif var.startswith('ma.'):
            ma_count += 1
    return ar_count, ma_count

def generar_campana_normal(residuos, media, std):
    if std == 0 or len(residuos) == 0:
        return [], []
    x = np.linspace(min(residuos), max(residuos), 100)
    y = (1 / (std * np.sqrt(2 * np.pi))) * np.exp(-0.5 * ((x - media) / std) ** 2)
    return x, y

def obtener_significancia_exogenas(coeficientes_df, exogenas_lista):
    if coeficientes_df is None or coeficientes_df.empty:
        return []
    resultados = []
    for exog in exogenas_lista:
        p_val = None
        for _, row in coeficientes_df.iterrows():
            var = str(row.get('Variable', ''))
            if exog in var:
                p_val = row.get('P_value', None)
                break
        if p_val is not None:
            try: p_val = float(p_val)
            except: p_val = None
        if p_val is None:
            resultados.append((exog, None, "Sin datos"))
        elif p_val < 0.05:
            resultados.append((exog, p_val, "Significativa"))
        elif p_val < 0.10:
            resultados.append((exog, p_val, "Marginal"))
        else:
            resultados.append((exog, p_val, "No significativa"))
    return resultados

def calcular_fwl_ponderado(fwl_df, pesos):
    if fwl_df is None or fwl_df.empty:
        return None
    fecha_col = None
    for c in fwl_df.columns:
        if 'fecha' in str(c).lower():
            fecha_col = c
            break
    base_col = [c for c in fwl_df.columns if 'FWL_BASE' in str(c).upper()]
    adv_col = [c for c in fwl_df.columns if 'FWL_ADVERSO' in str(c).upper() or 'FWL_ADVERSA' in str(c).upper()]
    opt_col = [c for c in fwl_df.columns if 'FWL_OPTIMISTA' in str(c).upper()]
    if not base_col or not adv_col or not opt_col:
        return None
    base_col, adv_col, opt_col = base_col[0], adv_col[0], opt_col[0]
    df = fwl_df.copy()
    df['FWL_Ponderado'] = (
        df[base_col].astype(float) * pesos['base'] +
        df[adv_col].astype(float) * pesos['adverso'] +
        df[opt_col].astype(float) * pesos['optimista']
    )
    return df

def resumen_fwl(fwl_df):
    if fwl_df is None or fwl_df.empty or 'FWL_Ponderado' not in fwl_df.columns:
        return {}
    vals = fwl_df['FWL_Ponderado'].dropna().astype(float)
    if len(vals) == 0:
        return {}
    return {'promedio': vals.mean(), 'maximo': vals.max(), 'minimo': vals.min(), 'volatilidad': vals.std()}

# =============================================================================
# SCORE SYSTEM (A-D Grading)
# =============================================================================
def calcular_score(p_val, prueba_nombre=""):
    if p_val is None or pd.isna(p_val):
        return 'N/A', ESTADO_NEUTRAL
    try:
        p_val = float(p_val)
    except:
        return 'N/A', ESTADO_NEUTRAL
    prueba_lower = str(prueba_nombre).lower()
    if 'jarque' in prueba_lower or 'bera' in prueba_lower:
        if p_val > 0.10:
            return 'A', SCORE_COLORS['A']
        elif p_val > 0.05:
            return 'B', SCORE_COLORS['B']
        elif p_val > 0.01:
            return 'C', SCORE_COLORS['C']
        else:
            return 'D', SCORE_COLORS['D']
    else:
        if p_val > 0.10:
            return 'A', SCORE_COLORS['A']
        elif p_val > 0.05:
            return 'B', SCORE_COLORS['B']
        elif p_val > 0.01:
            return 'C', SCORE_COLORS['C']
        else:
            return 'D', SCORE_COLORS['D']

def obtener_score_prueba(pruebas_df, nombre_prueba):
    if pruebas_df is None or pruebas_df.empty:
        return 'N/A', None
    nombre_lower = nombre_prueba.lower()
    for _, row in pruebas_df.iterrows():
        prueba = str(row.get('Prueba', '')).lower()
        p_val = row.get('P_value', None)
        if 'ljung' in nombre_lower and ('ljung' in prueba or 'box' in prueba):
            return calcular_score(p_val, prueba)
        elif 'jarque' in nombre_lower and ('jarque' in prueba or 'bera' in prueba):
            return calcular_score(p_val, prueba)
        elif 'hetero' in nombre_lower and ('hetero' in prueba or 'arch' in prueba):
            return calcular_score(p_val, prueba)
    return 'N/A', None

def obtener_scores_modelo(pruebas_df):
    scores = {}
    scores['ljung_box'] = obtener_score_prueba(pruebas_df, 'ljung')
    scores['jarque_bera'] = obtener_score_prueba(pruebas_df, 'jarque')
    scores['heterocedasticidad'] = obtener_score_prueba(pruebas_df, 'hetero')
    return scores

def calcular_score_global(pruebas_df):
    scores = obtener_scores_modelo(pruebas_df)
    acumulado, peso_total = 0.0, 0.0
    detalle = {}
    for clave, peso in PESOS_SCORE_GLOBAL.items():
        letra, _ = scores.get(clave, ('N/A', None))
        detalle[clave] = letra
        if letra in SCORE_NUM_MAP:
            acumulado += SCORE_NUM_MAP[letra] * peso
            peso_total += peso
    if peso_total == 0:
        return None, detalle
    return round(acumulado / peso_total, 1), detalle

def clasificar_score_global(score_num):
    if score_num is None:
        return "N/A", "No hay informacion suficiente para concluir la calidad global del modelo."
    if score_num >= 7:
        return "BUENO", "El modelo presenta diagnosticos robustos y consistentes para uso oficial."
    if score_num >= 5:
        return "REGULAR", "El modelo es util, pero requiere seguimiento y ajustes focalizados."
    return "DEFICIENTE", "El modelo no cumple criterios minimos para seleccion final sin reprocesamiento."

def estilo_score_global(score_num):
    etiqueta, _ = clasificar_score_global(score_num)
    if etiqueta == "N/A":
        return etiqueta, GRAY, ESTADO_NEUTRAL[0]
    if etiqueta == "BUENO":
        return etiqueta, GREEN, ESTADO_OK[0]
    if etiqueta == "REGULAR":
        return etiqueta, "#B8860B", ESTADO_WARN[0]
    return etiqueta, RED, ESTADO_FAIL[0]

def score_global_badge(score, tamano="12px"):
    etiqueta, color, bg = estilo_score_global(score)
    valor = f"{score:.1f}/10" if score is not None else "N/A"
    return (f'<span style="background:{bg};color:{color};font-size:{tamano};padding:3px 10px;'
            f'border-radius:4px;font-weight:700;">{valor} - {etiqueta}</span>')

def texto_ljungbox(score, p):
    p_str = fmt_pvalor(p)
    if score == "A":
        return f"A: p={p_str}. Sin evidencia de autocorrelacion en residuos."
    if score == "B":
        return f"B: p={p_str}. Comportamiento aceptable, sin evidencia fuerte de autocorrelacion."
    if score == "C":
        return f"C: p={p_str}. Senal moderada de autocorrelacion; requiere monitoreo."
    if score == "D":
        return f"D: p={p_str}. Evidencia clara de autocorrelacion; requiere ajuste del modelo."
    return "Sin datos para interpretar Ljung-Box."

def texto_jarquebera(score, p):
    p_str = fmt_pvalor(p)
    if score == "A":
        return f"A: p={p_str}. Residuos consistentes con normalidad."
    if score == "B":
        return f"B: p={p_str}. Normalidad razonable para analisis operativo."
    if score == "C":
        return f"C: p={p_str}. Posible desvio de normalidad; revisar outliers o transformaciones."
    if score == "D":
        return f"D: p={p_str}. No normalidad marcada de residuos."
    return "Sin datos para interpretar Jarque-Bera."

def texto_hetero(score, p):
    p_str = fmt_pvalor(p)
    if score == "A":
        return f"A: p={p_str}. Varianza de residuos estable."
    if score == "B":
        return f"B: p={p_str}. Sin evidencia significativa de heterocedasticidad."
    if score == "C":
        return f"C: p={p_str}. Senal debil de heterocedasticidad; monitorear estabilidad."
    if score == "D":
        return f"D: p={p_str}. Evidencia de heterocedasticidad; revisar especificacion."
    return "Sin datos para interpretar heterocedasticidad."

def interpretar_prueba(nombre_prueba, p_val, score):
    if score == 'N/A' or p_val is None or (isinstance(p_val, float) and pd.isna(p_val)):
        return "Sin datos disponibles para esta prueba."
    p_str = fmt_pvalor(p_val)
    nl = str(nombre_prueba).lower()
    if 'ljung' in nl or 'box' in nl:
        base = "autocorrelacion en los residuos"
    elif 'jarque' in nl or 'bera' in nl:
        base = "no-normalidad en los residuos"
    else:
        base = "heterocedasticidad (varianza no constante)"
    if score in ['A', 'B']:
        return f"p = {p_str} - sin evidencia significativa de {base}."
    elif score == 'C':
        return f"p = {p_str} - evidencia debil de {base}."
    else:
        return f"p = {p_str} - evidencia de {base}."

def score_badge(score, bg_color, fg_color):
    return f'<span style="background:{bg_color};color:{fg_color};font-size:12px;padding:3px 10px;border-radius:4px;font-weight:700;">{score}</span>'

def render_leyenda_scores():
    html = f"""
    <div style="background:{WHITE};border:1px solid {BORDER};border-radius:8px;padding:16px;margin:16px 0;">
        <p style="font-size:12px;font-weight:700;color:{NAVY};margin:0 0 12px;text-transform:uppercase;letter-spacing:0.5px;">Clasificacion de Scores</p>
        <table style="width:100%;border-collapse:collapse;font-size:12px;">
            <thead>
                <tr style="border-bottom:1px solid {BORDER};">
                    <th style="text-align:left;padding:6px 8px;font-weight:600;color:{NAVY};">Score</th>
                    <th style="text-align:left;padding:6px 8px;font-weight:600;color:{NAVY};">Rango p-valor</th>
                    <th style="text-align:left;padding:6px 8px;font-weight:600;color:{NAVY};">Ljung-Box / Heterocedasticidad</th>
                    <th style="text-align:left;padding:6px 8px;font-weight:600;color:{NAVY};">Jarque-Bera</th>
                </tr>
            </thead>
            <tbody>
                <tr style="border-bottom:1px solid {BORDER};">
                    <td style="padding:6px 8px;">{score_badge('A', SCORE_COLORS['A'][0], SCORE_COLORS['A'][1])}</td>
                    <td style="padding:6px 8px;color:{TEXT};">p &gt; 0.10</td>
                    <td style="padding:6px 8px;color:{GREEN};font-weight:600;">Sin autocorrelacion / heterocedasticidad</td>
                    <td style="padding:6px 8px;color:{GREEN};font-weight:600;">Residuos normales</td>
                </tr>
                <tr style="border-bottom:1px solid {BORDER};">
                    <td style="padding:6px 8px;">{score_badge('B', SCORE_COLORS['B'][0], SCORE_COLORS['B'][1])}</td>
                    <td style="padding:6px 8px;color:{TEXT};">0.05 &lt; p &le; 0.10</td>
                    <td style="padding:6px 8px;color:{BLUE};font-weight:600;">Sin evidencia significativa</td>
                    <td style="padding:6px 8px;color:{BLUE};font-weight:600;">Sin evidencia significativa</td>
                </tr>
                <tr style="border-bottom:1px solid {BORDER};">
                    <td style="padding:6px 8px;">{score_badge('C', SCORE_COLORS['C'][0], SCORE_COLORS['C'][1])}</td>
                    <td style="padding:6px 8px;color:{TEXT};">0.01 &lt; p &le; 0.05</td>
                    <td style="padding:6px 8px;color:#B8860B;font-weight:600;">Evidencia debil de problema</td>
                    <td style="padding:6px 8px;color:#B8860B;font-weight:600;">Evidencia debil de no-normalidad</td>
                </tr>
                <tr>
                    <td style="padding:6px 8px;">{score_badge('D', SCORE_COLORS['D'][0], SCORE_COLORS['D'][1])}</td>
                    <td style="padding:6px 8px;color:{TEXT};">p &le; 0.01</td>
                    <td style="padding:6px 8px;color:{RED};font-weight:600;">Autocorrelacion / heterocedasticidad presente</td>
                    <td style="padding:6px 8px;color:{RED};font-weight:600;">Residuos no normales</td>
                </tr>
            </tbody>
        </table>
        <p style="font-size:10px;color:{MUTED};margin:10px 0 0;line-height:1.4;">
            <b>Ljung-Box:</b> Prueba de autocorrelacion en residuos. H0: no hay autocorrelacion.<br>
            <b>Jarque-Bera:</b> Prueba de normalidad. H0: los residuos siguen distribucion normal.<br>
            <b>Heterocedasticidad (ARCH/LM):</b> Prueba de varianza constante. H0: varianza homogenea.
        </p>
    </div>
    """
    return html

# =============================================================================
# METADATA
# =============================================================================
def truncar_texto(texto, max_len=40):
    """Trunca texto largo con '...' para evitar desbordamiento en tablas PDF."""
    s = str(texto)
    return s if len(s) <= max_len else s[:max_len-3] + "..."

def extraer_kpis_meta(meta, fallback_meta=None):
    src = meta if meta else fallback_meta
    if not src:
        return {}
    return {
        'pais': PAISES_MAP.get(str(src.get('pais', '')).lower().strip(), src.get('pais', 'N/A')),
        'cartera': CARTERAS_MAP.get(str(src.get('cartera', '')).lower().strip(), src.get('cartera', 'N/A')),
        'tipo_endogena': src.get('motor_tipo_endogena', 'N/A'),
        'modo_endogena': src.get('generador_modo_endogena', 'N/A'),
        'ventana_mm': src.get('generador_ventana_mm', 'N/A'),
        'vif_max': src.get('motor_vif_max', 'N/A'),
        'fwl_min': src.get('motor_fwl_factor_min', '?'),
        'fwl_max': src.get('motor_fwl_factor_max', '?'),
        'max_exog': src.get('motor_max_exog_por_modelo', 'N/A'),
        'top_exportar': src.get('motor_top_exportar', 'N/A'),
    }

# =============================================================================
# DOCUMENTO METODOLOGICO MODELO FINAL
# =============================================================================
def _normalizar_nombre_archivo(texto):
    limpio = "".join(ch if str(ch).isalnum() or ch in ("_", "-", ".") else "_" for ch in str(texto))
    while "__" in limpio:
        limpio = limpio.replace("__", "_")
    return limpio.strip("_") or "modelo_final"

def _obtener_fecha_fin_hist(meta_contexto):
    if not meta_contexto:
        return None
    candidatos = [
        "generador_fecha_fin_hist", "fecha_fin_hist", "fecha_fin_historico",
        "motor_fecha_fin_hist", "FECHA_FIN_HIST"
    ]
    for clave in candidatos:
        if clave in meta_contexto and meta_contexto.get(clave):
            fecha = pd.to_datetime(meta_contexto.get(clave), errors="coerce")
            if not pd.isna(fecha):
                return fecha
    return None

def recolectar_datos_documento(nombre_modelo, modelos_data, meta_contexto):
    modelo = modelos_data.get(nombre_modelo, {})
    pruebas = modelo.get("pruebas")
    coeficientes = modelo.get("coeficientes")
    exogenas = modelo.get("exogenas_nombres", [])
    scores = obtener_scores_modelo(pruebas)
    score_global_num, detalle_global = calcular_score_global(pruebas)
    score_global_txt, score_global_conclusion = clasificar_score_global(score_global_num)
    significancia = obtener_significancia_exogenas(coeficientes, exogenas)
    ar_count, ma_count = contar_ar_ma(coeficientes)
    meta_kpis = extraer_kpis_meta(meta_contexto, st.session_state.get("meta_contexto_manual", {}))

    diag_rows = []
    if pruebas is not None and not pruebas.empty:
        for _, row in pruebas.iterrows():
            prueba = limpiar_nombre_prueba(row.get("Prueba", ""))
            p_val = row.get("P_value", None)
            score, _ = calcular_score(p_val, prueba)
            prueba_lower = str(prueba).lower()
            if "ljung" in prueba_lower or "box" in prueba_lower:
                interpretacion = texto_ljungbox(score, p_val)
            elif "jarque" in prueba_lower or "bera" in prueba_lower:
                interpretacion = texto_jarquebera(score, p_val)
            else:
                interpretacion = texto_hetero(score, p_val)
            diag_rows.append({
                "prueba": prueba,
                "estadistico": row.get("Estadistico"),
                "p_value": p_val,
                "score": score,
                "interpretacion": interpretacion,
            })

    exog_rows = [
        {"exogena": ex, "p_value": p_val, "estado": estado}
        for ex, p_val, estado in significancia
    ]

    coefs_rows = []
    if coeficientes is not None and not coeficientes.empty:
        for _, row in coeficientes.iterrows():
            coefs_rows.append({
                "variable": row.get("Variable"),
                "coeficiente": row.get("Coeficiente"),
                "p_value": row.get("P_value"),
                "tipo": clasificar_variable(row.get("Variable", "")),
            })

    return {
        "nombre_modelo": nombre_modelo,
        "fecha_generacion": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "meta_contexto": meta_contexto or {},
        "meta_kpis": meta_kpis,
        "score_global_num": score_global_num,
        "score_global_detalle": detalle_global,
        "score_global_clase": score_global_txt,
        "score_global_conclusion": score_global_conclusion,
        "scores": scores,
        "diagnosticos": diag_rows,
        "exogenas": exog_rows,
        "coeficientes": coefs_rows,
        "ar_count": ar_count,
        "ma_count": ma_count,
        "fwl_min": meta_kpis.get("fwl_min", "N/A"),
        "fwl_max": meta_kpis.get("fwl_max", "N/A"),
        "max_exog": meta_kpis.get("max_exog", "N/A"),
        "vif_max": meta_kpis.get("vif_max", "N/A"),
        "top_exportar": meta_kpis.get("top_exportar", "N/A"),
        "umbral_sensibilidad": (meta_contexto or {}).get("motor_umbral_sensibilidad", "N/A"),
        "data_exogenas": modelo.get("exogenas"),
        "data_endogena": modelo.get("fecha_endogena"),
        "endogenas_cols": modelo.get("endogenas_cols", ["BASE", "ADVERSO", "OPTIMISTA"]),
        "observaciones": modelo.get("observaciones", 0),
        "data_fwl_12m": modelo.get("fwl_12m"),
        "fecha_fin_hist": _obtener_fecha_fin_hist(meta_contexto or {}),
    }

def graficar_macros_estatico(doc_data, output_path):
    df_exog = doc_data.get("data_exogenas")
    if df_exog is None or df_exog.empty:
        return None
    df = df_exog.copy()
    fecha_col = "fecha" if "fecha" in df.columns else df.columns[0]
    df[fecha_col] = pd.to_datetime(df[fecha_col], errors="coerce")
    df = df.dropna(subset=[fecha_col]).sort_values(fecha_col)
    if df.empty:
        return None

    # Detectar variables exógenas base (sin sufijo _BASE/_ADVERSO/_OPTIMISTA)
    escenarios = ["_BASE", "_ADVERSO", "_OPTIMISTA"]
    # Paleta corporativa del app (misma que Base/Adverso/Optimista en Plotly)
    colores_esc = {"_BASE": BLUE, "_ADVERSO": RED, "_OPTIMISTA": GREEN}
    nombres_esc = {"_BASE": "Base", "_ADVERSO": "Adverso", "_OPTIMISTA": "Optimista"}

    # Identificar variables base únicas
    vars_base = set()
    for col in df.columns:
        if col == fecha_col:
            continue
        base_name = col
        for suf in escenarios:
            if str(col).upper().endswith(suf):
                base_name = col[:-len(suf)]
                break
        vars_base.add(base_name)

    vars_base = sorted([v for v in vars_base if v != fecha_col])
    if not vars_base:
        return None

    n_vars = len(vars_base)

    # Una variable por fila (apiladas verticalmente) para que se vean bien en el PDF
    n_cols = 1
    n_rows = n_vars

    plt.close("all")
    fig, axes = plt.subplots(n_rows, n_cols, figsize=(10, 3.3 * n_rows), squeeze=False)
    axes = axes.flatten()

    fecha_fin_hist = doc_data.get("fecha_fin_hist")

    for idx, var_base in enumerate(vars_base):
        ax = axes[idx]
        for suf in escenarios:
            col_name = var_base + suf
            col_match = None
            for c in df.columns:
                if str(c).upper() == col_name.upper():
                    col_match = c
                    break
            if col_match is not None:
                serie = pd.to_numeric(df[col_match], errors="coerce")
                if serie.notna().any():
                    ax.plot(
                        df[fecha_col], serie,
                        label=nombres_esc[suf],
                        color=colores_esc[suf],
                        linewidth=1.8
                    )

        if fecha_fin_hist is not None and not pd.isna(fecha_fin_hist):
            ax.axvline(
                pd.to_datetime(fecha_fin_hist),
                color=GRAY, linestyle="--", linewidth=1, label="Fin histórico"
            )

        ax.set_facecolor(WHITE)
        ax.set_title(var_base, fontsize=11, fontweight="bold", color=NAVY)
        ax.set_xlabel("Fecha", fontsize=8, color=MUTED)
        ax.set_ylabel("Valor", fontsize=8, color=MUTED)
        ax.grid(alpha=0.3, color=BORDER)
        ax.legend(fontsize=7, loc="best", frameon=False)
        ax.tick_params(axis="both", labelsize=7, colors=MUTED)
        for spine in ax.spines.values():
            spine.set_color(BORDER)

    # Ocultar ejes sobrantes
    for idx in range(n_vars, len(axes)):
        axes[idx].set_visible(False)

    fig.patch.set_facecolor(WHITE)
    plt.tight_layout()
    try:
        plt.savefig(output_path, dpi=150, bbox_inches="tight", facecolor=WHITE)
    finally:
        plt.close("all")
    return output_path


def graficar_fwl_estatico(doc_data, output_path):
    """Genera gráfica del Factor FWL a 12 meses con matplotlib."""
    df_fwl = doc_data.get("data_fwl_12m")
    if df_fwl is None or df_fwl.empty:
        return None
    df = df_fwl.copy()
    fecha_col = None
    for c in df.columns:
        if 'fecha' in str(c).lower():
            fecha_col = c
            break
    if fecha_col is None:
        fecha_col = df.columns[0]
    df[fecha_col] = pd.to_datetime(df[fecha_col], errors="coerce")
    df = df.dropna(subset=[fecha_col]).sort_values(fecha_col)
    if df.empty:
        return None

    plt.close("all")
    fig, ax = plt.subplots(figsize=(10, 4.5))

    colores = {"FWL_BASE": "#2a7f3f", "FWL_ADVERSO": "#b22222", "FWL_OPTIMISTA": "#e08a00"}
    nombres = {"FWL_BASE": "Base", "FWL_ADVERSO": "Adverso", "FWL_OPTIMISTA": "Optimista"}

    for col in df.columns:
        col_up = str(col).upper()
        if col_up in colores:
            serie = pd.to_numeric(df[col], errors="coerce")
            if serie.notna().any():
                ax.plot(df[fecha_col], serie, label=nombres.get(col_up, col), color=colores[col_up], linewidth=2)

    fecha_fin_hist = doc_data.get("fecha_fin_hist")
    if fecha_fin_hist is not None and not pd.isna(fecha_fin_hist):
        ax.axvline(pd.to_datetime(fecha_fin_hist), color="#555555", linestyle="--", linewidth=1.5, label="Fin histórico")

    ax.set_title("Factor FWL a 12 Meses", fontsize=12, fontweight="bold", pad=10)
    ax.set_xlabel("Fecha", fontsize=9)
    ax.set_ylabel("FWL", fontsize=9)
    ax.grid(alpha=0.25)
    ax.legend(fontsize=8, loc="best")
    ax.tick_params(axis="both", labelsize=8)
    plt.tight_layout()
    try:
        plt.savefig(output_path, dpi=150, bbox_inches="tight")
    finally:
        plt.close("all")
    return output_path

def graficar_historico_estatico(doc_data, output_path):
    """Genera gráfica de la endógena histórica/proyectada con matplotlib."""
    df_end = doc_data.get("data_endogena")
    endogena_cols = doc_data.get("endogenas_cols", ["BASE", "ADVERSO", "OPTIMISTA"])
    if df_end is None or df_end.empty:
        return None
    df = df_end.copy()
    fecha_col = "fecha" if "fecha" in df.columns else df.columns[0]
    df[fecha_col] = pd.to_datetime(df[fecha_col], errors="coerce")
    df = df.dropna(subset=[fecha_col]).sort_values(fecha_col)
    if df.empty:
        return None

    plt.close("all")
    fig, ax = plt.subplots(figsize=(10, 4.5))

    colores = {"BASE": BLUE, "ADVERSO": RED, "OPTIMISTA": GREEN}
    nombres = {"BASE": "Base", "ADVERSO": "Adverso", "OPTIMISTA": "Optimista"}

    for col in endogena_cols:
        col_up = str(col).upper()
        if col_up in ["ADVERSO", "ADVERSA"]:
            col_up = "ADVERSO"
        if col in df.columns:
            serie = pd.to_numeric(df[col], errors="coerce")
            if serie.notna().any():
                ax.plot(
                    df[fecha_col], serie,
                    label=nombres.get(col_up, col),
                    color=colores.get(col_up, GRAY),
                    linewidth=2
                )

    fecha_fin_hist = doc_data.get("fecha_fin_hist")
    if fecha_fin_hist is not None and not pd.isna(fecha_fin_hist):
        ax.axvline(
            pd.to_datetime(fecha_fin_hist),
            color=GRAY, linestyle="--", linewidth=1.5, label="Fin histórico"
        )

    ax.set_title("Evolución Histórica y Proyectada", fontsize=12, fontweight="bold", pad=10)
    ax.set_xlabel("Fecha", fontsize=9)
    ax.set_ylabel("Valor", fontsize=9)
    ax.grid(alpha=0.25)
    ax.legend(fontsize=8, loc="best")
    ax.tick_params(axis="both", labelsize=8)
    plt.tight_layout()
    try:
        plt.savefig(output_path, dpi=150, bbox_inches="tight")
    finally:
        plt.close("all")
    return output_path

def texto_ljungbox(score: str, p: float) -> str:
    if score == "A":
        return (
            "No se detecta autocorrelación significativa en los residuos (p > 0.10). El "
            "modelo ha capturado adecuadamente la estructura temporal de la serie, por lo "
            "que no hay evidencia de patrones no explicados en los errores de predicción. "
            "Esto valida que los componentes AR y MA seleccionados son suficientes para "
            "modelar la dependencia temporal."
        )
    elif score == "B":
        return (
            "No existe evidencia concluyente de autocorrelación residual (0.05 < p ≤ 0.10). "
            "Aunque el estadístico se acerca al umbral de significancia, el modelo "
            "mantiene un comportamiento aceptable y no se requieren ajustes estructurales "
            "inmediatos. Se recomienda monitorear en corridas posteriores."
        )
    elif score == "C":
        return (
            "Se detecta evidencia débil de autocorrelación en los residuos (p ≤ 0.05). "
            "Esto sugiere que podría existir una estructura temporal no capturada por el "
            "modelo actual. Se recomienda evaluar la inclusión de rezagos adicionales o "
            "revisar la especificación ARMA (aumentar p o q)."
        )
    else:  # D
        return (
            "Se detecta autocorrelación significativa en los residuos (p ≤ 0.01). El "
            "modelo NO está capturando adecuadamente la dinámica temporal de la serie, lo "
            "que puede generar predicciones sesgadas e intervalos de confianza inválidos. "
            "Se requiere reespecificación urgente del modelo (revisar órdenes AR/MA o "
            "considerar estacionalidad)."
        )

def texto_jarquebera(score: str, p: float) -> str:
    if score == "A":
        return (
            "Los residuos siguen una distribución normal (p > 0.10). Esto valida el uso "
            "de intervalos de confianza estándar y garantiza la validez de las inferencias "
            "estadísticas derivadas del modelo. Los errores no presentan asimetría ni "
            "curtosis problemáticas."
        )
    elif score == "B":
        return (
            "No se encuentra evidencia concluyente de no-normalidad (0.05 < p ≤ 0.10). "
            "La distribución de los errores es suficientemente simétrica para mantener la "
            "confiabilidad de las proyecciones, aunque se sugiere monitoreo periódico."
        )
    elif score == "C":
        return (
            "Se detecta evidencia débil de no-normalidad en los residuos (p ≤ 0.05). La "
            "presencia de asimetría o curtosis atípica puede afectar la precisión de los "
            "intervalos de predicción extremos. Se sugiere monitorear especialmente los "
            "escenarios adverso y optimista."
        )
    else:  # D
        return (
            "Los residuos presentan desviaciones significativas de la normalidad (p ≤ 0.01): "
            "asimetría y/o curtosis extrema. Esto invalida los intervalos de confianza "
            "clásicos y puede indicar la presencia de valores atípicos no modelados o una "
            "especificación incorrecta. Se recomienda transformar la variable (ej. logit) "
            "o utilizar métodos robustos."
        )

def texto_hetero(score: str, p: float) -> str:
    if score == "A":
        return (
            "No se detecta heterocedasticidad (p > 0.10). La varianza de los residuos es "
            "constante a lo largo del tiempo, cumpliendo con el supuesto de homocedasticidad "
            "requerido para la inferencia válida del modelo. Los errores estándar de los "
            "coeficientes son confiables."
        )
    elif score == "B":
        return (
            "No existe evidencia concluyente de varianza no constante (0.05 < p ≤ 0.10). "
            "El modelo mantiene estabilidad en la dispersión de los errores a través de la "
            "muestra. No se requieren correcciones adicionales."
        )
    elif score == "C":
        return (
            "Se detecta evidencia débil de heterocedasticidad (p ≤ 0.05). La varianza de "
            "los errores presenta cambios moderados en el tiempo, lo que puede afectar la "
            "eficiencia de los estimadores. Se sugiere considerar modelos ARCH/GARCH para "
            "la volatilidad o utilizar errores estándar robustos (White)."
        )
    else:  # D
        return (
            "Se confirma heterocedasticidad significativa (p ≤ 0.01). La varianza de los "
            "residuos NO es constante, lo que genera estimadores ineficientes y errores "
            "estándar sesgados. Las proyecciones pueden subestimar o sobrestimar el riesgo "
            "en diferentes períodos. Se requiere corrección: errores estándar robustos o "
            "modelos de varianza condicional (ARCH/GARCH)."
        )

def conclusion_score_global_pdf(s: float) -> str:
    """Retorna solo la conclusión larga para el PDF."""
    if s >= 7:
        return (
            "El modelo presenta un desempeño general satisfactorio. Los diagnósticos "
            "estadísticos indican que la especificación captura adecuadamente la dinámica "
            "de la cartera y las proyecciones son confiables para la estimación de provisiones "
            "bajo IFRS 9. Se recomienda su uso como modelo oficial de proyección."
        )
    elif s >= 5:
        return (
            "El modelo presenta un desempeño aceptable con áreas de mejora identificadas. "
            "Aunque las proyecciones son utilizables, se recomienda monitorear los diagnósticos "
            "marcados con score C o D en corridas posteriores y evaluar reespecificaciones "
            "si la calidad del ajuste se deteriora."
        )
    else:
        return (
            "El modelo presenta problemas estadísticos significativos que comprometen la "
            "confiabilidad de las proyecciones. Se recomienda NO utilizar este modelo como "
            "salida oficial sin una reespecificación profunda que aborde los diagnósticos "
            "con score D y, de ser posible, los score C."
        )


# =============================================================================
# GENERADOR DE PDF CON REPORTLAB
# =============================================================================

def generar_pdf_metodologico(doc_data: dict, ruta_imagen_exog: str, ruta_imagen_fwl: str = None, ruta_imagen_historico: str = None) -> bytes:
    """
    Genera el documento metodológico completo en PDF usando reportlab.
    Consume la estructura exacta devuelta por recolectar_datos_documento().
    """
    import io
    import os
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        Image, PageBreak
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
    from datetime import datetime

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=72,
        leftMargin=72,
        topMargin=72,
        bottomMargin=72
    )

    styles = getSampleStyleSheet()

    # Paleta corporativa (misma que el dashboard: NAVY/BLUE/GREEN/RED)
    NAVY_CORP = colors.HexColor(NAVY)
    AZUL_CORP = colors.HexColor(BLUE)
    VERDE_CORP = colors.HexColor(GREEN)
    ROJO_CORP = colors.HexColor(RED)
    AMBAR_CORP = colors.HexColor("#B8860B")   # mismo tono que Score C / REGULAR en el app
    GRIS_OSCURO = colors.HexColor(TEXT)
    GRIS_CLARO = colors.HexColor(TINT)
    BLANCO = colors.white

    estilo_titulo = ParagraphStyle(
        "TituloDoc",
        parent=styles["Heading1"],
        fontSize=22,
        leading=26,
        textColor=NAVY_CORP,
        alignment=TA_CENTER,
        spaceAfter=20,
        fontName="Helvetica-Bold"
    )

    estilo_subtitulo = ParagraphStyle(
        "SubtituloDoc",
        parent=styles["Heading2"],
        fontSize=14,
        leading=18,
        textColor=GRIS_OSCURO,
        alignment=TA_CENTER,
        spaceAfter=30,
        fontName="Helvetica"
    )

    estilo_seccion = ParagraphStyle(
        "SeccionDoc",
        parent=styles["Heading2"],
        fontSize=16,
        leading=20,
        textColor=NAVY_CORP,
        spaceBefore=20,
        spaceAfter=12,
        fontName="Helvetica-Bold",
        leftIndent=0
    )

    estilo_subseccion = ParagraphStyle(
        "SubseccionDoc",
        parent=styles["Heading3"],
        fontSize=12,
        leading=15,
        textColor=GRIS_OSCURO,
        spaceBefore=12,
        spaceAfter=8,
        fontName="Helvetica-Bold"
    )

    estilo_cuerpo = ParagraphStyle(
        "CuerpoDoc",
        parent=styles["BodyText"],
        fontSize=10,
        leading=14,
        textColor=GRIS_OSCURO,
        alignment=TA_JUSTIFY,
        spaceAfter=10,
        fontName="Helvetica"
    )

    estilo_nota = ParagraphStyle(
        "NotaDoc",
        parent=styles["BodyText"],
        fontSize=9,
        leading=12,
        textColor=colors.grey,
        alignment=TA_LEFT,
        spaceAfter=8,
        fontName="Helvetica-Oblique"
    )

    estilo_score_a = ParagraphStyle("ScoreA", parent=estilo_cuerpo, textColor=VERDE_CORP, fontName="Helvetica-Bold")
    estilo_score_b = ParagraphStyle("ScoreB", parent=estilo_cuerpo, textColor=AZUL_CORP, fontName="Helvetica-Bold")
    estilo_score_c = ParagraphStyle("ScoreC", parent=estilo_cuerpo, textColor=AMBAR_CORP, fontName="Helvetica-Bold")
    estilo_score_d = ParagraphStyle("ScoreD", parent=estilo_cuerpo, textColor=ROJO_CORP, fontName="Helvetica-Bold")

    def estilo_por_score(score: str):
        return {"A": estilo_score_a, "B": estilo_score_b, "C": estilo_score_c, "D": estilo_score_d}.get(score, estilo_cuerpo)

    def tabla_estilo_base():
        return TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), NAVY_CORP),
            ("TEXTCOLOR", (0, 0), (-1, 0), BLANCO),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
            ("BACKGROUND", (0, 1), (-1, -1), GRIS_CLARO),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(BORDER)),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ])

    story = []

    # ------------------------------------------------------------------
    # Extraer datos de doc_data
    # ------------------------------------------------------------------
    nombre_modelo = doc_data.get("nombre_modelo", "N/A")
    meta_kpis = doc_data.get("meta_kpis", {})
    meta_ctx = doc_data.get("meta_contexto", {})
    score_global = doc_data.get("score_global_num")
    if score_global is None:
        score_global = 0.0
    clasificacion = doc_data.get("score_global_clase", "N/A")
    conclusion_global = doc_data.get("score_global_conclusion", "")
    if not conclusion_global:
        conclusion_global = conclusion_score_global_pdf(score_global)
    diagnosticos = doc_data.get("diagnosticos", [])
    coeficientes = doc_data.get("coeficientes", [])
    exogenas_list = doc_data.get("exogenas", [])
    ar_count = doc_data.get("ar_count", 0)
    ma_count = doc_data.get("ma_count", 0)
    observaciones = doc_data.get("observaciones", 0)
    fecha_gen = doc_data.get("fecha_generacion", datetime.now().strftime("%d/%m/%Y %H:%M"))

    meta = {
        "pais": meta_kpis.get("pais", "N/A"),
        "cartera": meta_kpis.get("cartera", "N/A"),
        "tipo_endogena": meta_kpis.get("tipo_endogena", "N/A"),
        "modo_endogena": meta_kpis.get("modo_endogena", "N/A"),
        "ventana_mm": meta_kpis.get("ventana_mm", "N/A"),
        "vif_max": meta_kpis.get("vif_max", "N/A"),
        "max_exog": meta_kpis.get("max_exog", "N/A"),
        "top_exportar": meta_kpis.get("top_exportar", "N/A"),
        "fwl_min": meta_kpis.get("fwl_min", "N/A"),
        "fwl_max": meta_kpis.get("fwl_max", "N/A"),
        "valores_p": meta_ctx.get("motor_valores_p", "N/A"),
        "valores_q": meta_ctx.get("motor_valores_q", "N/A"),
        "max_lags": meta_ctx.get("motor_max_lags", "N/A"),
        "umbral_sensibilidad": doc_data.get("umbral_sensibilidad", meta_ctx.get("motor_umbral_sensibilidad", "N/A")),
        "trend_modelo": meta_ctx.get("motor_trend", "N/A"),
    }

    # Buscar diagnósticos por nombre
    diag_map = {}
    for d in diagnosticos:
        nombre_prueba = str(d.get("prueba", "")).lower()
        if "ljung" in nombre_prueba or "box" in nombre_prueba:
            diag_map["ljung_box"] = d
        elif "jarque" in nombre_prueba or "bera" in nombre_prueba:
            diag_map["jarque_bera"] = d
        elif "hetero" in nombre_prueba or "arch" in nombre_prueba:
            diag_map["heterocedasticidad"] = d

    lb = diag_map.get("ljung_box", {})
    jb = diag_map.get("jarque_bera", {})
    ht = diag_map.get("heterocedasticidad", {})

    def fmt_num(v):
        try:
            return f"{float(v):.4f}"
        except:
            return str(v)

    # =====================================================================
    # PORTADA
    # =====================================================================
    story.append(Spacer(1, 1.5 * inch))
    story.append(Paragraph("Documento Metodológico", estilo_titulo))
    story.append(Paragraph("Modelo SARIMAX — Proyección IFRS 9", estilo_subtitulo))
    story.append(Spacer(1, 0.3 * inch))

    datos_portada = [
        ["Modelo seleccionado:", nombre_modelo],
        ["Fecha de generación:", fecha_gen],
        ["País:", meta.get("pais", "N/A")],
        ["Cartera:", meta.get("cartera", "N/A")],
    ]
    if score_global is not None and score_global >= 7:
        datos_portada.append(["Score global:", f"{score_global:.2f} / 10 — {clasificacion}"])
    tabla_portada = Table(datos_portada, colWidths=[2.2 * inch, 3.5 * inch])
    tabla_portada.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), GRIS_CLARO),
        ("TEXTCOLOR", (0, 0), (-1, -1), GRIS_OSCURO),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("ALIGN", (1, 0), (1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(tabla_portada)
    story.append(PageBreak())

    # =====================================================================
    # SECCIÓN 0 — RESUMEN DEL MODELO
    # =====================================================================
    story.append(Paragraph("Resumen del Modelo", estilo_seccion))
    story.append(Paragraph(
        "Esta sección presenta una vista consolidada del modelo seleccionado, incluyendo el factor FWL, "
        "los diagnósticos estadísticos, los coeficientes estimados y la estructura del modelo.",
        estilo_cuerpo
    ))
    story.append(Spacer(1, 0.1 * inch))

    # 0.0 Evolución histórica y proyectada de la endógena
    story.append(Paragraph("Evolución Histórica y Proyectada", estilo_subseccion))
    if ruta_imagen_historico and os.path.exists(ruta_imagen_historico):
        try:
            from PIL import Image as PILImage
            with PILImage.open(ruta_imagen_historico) as im:
                img_w_px, img_h_px = im.size
            aspecto = img_h_px / img_w_px
            ancho_max = 6.5 * inch
            alto_max = 4.5 * inch
            ancho = ancho_max
            alto = ancho * aspecto
            if alto > alto_max:
                alto = alto_max
                ancho = alto / aspecto
            story.append(Image(ruta_imagen_historico, width=ancho, height=alto))
        except Exception:
            story.append(Image(ruta_imagen_historico, width=6.5 * inch, height=4 * inch))
    else:
        story.append(Paragraph("[Gráfica histórica no disponible]", estilo_nota))
    story.append(Spacer(1, 0.15 * inch))

    # 0.1 Factor FWL a 12 meses
    story.append(Paragraph("Factor FWL a 12 Meses", estilo_subseccion))
    if ruta_imagen_fwl and os.path.exists(ruta_imagen_fwl):
        img_fwl = Image(ruta_imagen_fwl, width=6.5 * inch, height=3 * inch)
        story.append(img_fwl)
    else:
        story.append(Paragraph("[Gráfica FWL no disponible]", estilo_nota))
    story.append(Spacer(1, 0.15 * inch))

    # 0.2 Diagnósticos resumidos
    story.append(Paragraph("Diagnósticos Estadísticos", estilo_subseccion))
    diag_resumen = [
        ["Prueba", "Score", "p-valor", "Estadístico"],
        ["Ljung-Box", lb.get("score", "N/A"), fmt_num(lb.get("p_value")), fmt_num(lb.get("estadistico"))],
        ["Jarque-Bera", jb.get("score", "N/A"), fmt_num(jb.get("p_value")), fmt_num(jb.get("estadistico"))],
        ["Heterocedasticidad", ht.get("score", "N/A"), fmt_num(ht.get("p_value")), fmt_num(ht.get("estadistico"))],
    ]
    if score_global is not None and score_global >= 7:
        diag_resumen.append(["Score Global Ponderado", "", "", f"{score_global:.2f} / 10"])
    tabla_diag = Table(diag_resumen, colWidths=[2.2 * inch, 1.0 * inch, 1.2 * inch, 1.3 * inch])
    tabla_diag.setStyle(tabla_estilo_base())
    story.append(tabla_diag)
    story.append(Spacer(1, 0.15 * inch))

    # 0.3 Coeficientes del modelo
    story.append(Paragraph("Coeficientes del Modelo", estilo_subseccion))
    coef_data = [["Variable", "Coeficiente", "p-value", "Tipo"]]
    coef_count = 0
    for row in coeficientes:
        coef_count += 1
        pval = row.get("p_value")
        coef_data.append([
            str(row.get("variable", "N/A")),
            f"{float(row.get('coeficiente', 0)):.6f}" if row.get('coeficiente') is not None else "N/A",
            fmt_num(pval),
            str(row.get("tipo", "N/A"))
        ])
    if coef_count > 0:
        # Truncar nombres de variables muy largos para evitar desbordamiento
        for row in coef_data[1:]:
            row[0] = truncar_texto(row[0], max_len=42)
        tabla_coef = Table(coef_data, colWidths=[3.2 * inch, 1.1 * inch, 1.0 * inch, 0.9 * inch])
        tabla_coef.setStyle(tabla_estilo_base())
        story.append(tabla_coef)
    else:
        story.append(Paragraph("No hay datos de coeficientes.", estilo_cuerpo))
    story.append(Spacer(1, 0.15 * inch))

    # 0.4 Estructura del modelo
    story.append(Paragraph("Estructura del Modelo", estilo_subseccion))
    exo_sig_count = sum(1 for row in coeficientes if row.get("tipo") == "Exogena")
    struct_data = [
        ["Componente", "Valor"],
        ["Términos AR", str(ar_count)],
        ["Términos MA", str(ma_count)],
        ["Variables exógenas", str(exo_sig_count)],
        ["Observaciones", str(observaciones)],
    ]
    if score_global is not None and score_global >= 7:
        struct_data.append(["Score global", f"{score_global:.2f} / 10 — {clasificacion}"])
    tabla_struct = Table(struct_data, colWidths=[2.5 * inch, 3.2 * inch])
    tabla_struct.setStyle(tabla_estilo_base())
    story.append(tabla_struct)
    story.append(PageBreak())

    # =====================================================================
    # SECCIÓN 1 — METODOLOGÍA SARIMAX (separada por bloques)
    # =====================================================================
    story.append(Paragraph("1. Metodología SARIMAX", estilo_seccion))

    story.append(Paragraph(
        "El modelo SARIMAX (Seasonal AutoRegressive Integrated Moving Average with eXogenous regressors) "
        "es una extensión del modelo ARIMA que incorpora componentes estacionales y variables exógenas. "
        "En el contexto de IFRS 9, se utiliza para proyectar la variación de la cartera de crédito bajo diferentes "
        "escenarios macroeconómicos (base, adverso y optimista). El modelo descompone la serie histórica en "
        "componentes autorregresivos (AR), de medias móviles (MA), diferenciación integrada (I) y efectos "
        "estacionales (S), permitiendo capturar la dinámica temporal inherente a los datos. La inclusión de "
        "variables exógenas (macroeconómicas y administrativas) permite vincular el comportamiento de la "
        "cartera con el entorno económico, cumpliendo con el requisito de forward-looking information establecido "
        "en la norma. El flujo se organiza en nueve bloques secuenciales, desde la carga de datos hasta la exportación "
        "de resultados.",
        estilo_cuerpo
    ))
    story.append(Spacer(1, 0.1 * inch))

    story.append(Paragraph("1.1 Identificación de Variables", estilo_subseccion))
    story.append(Paragraph(
        "El proceso inicia con la carga de cuatro archivos base: (1) ENDÓGENA, que contiene la fecha, la variable "
        "dependiente y las n exógenas históricas; (2) EXO_BAS, con las proyecciones base de las exógenas; "
        "(3) EXO_ADV, con las proyecciones adversas; y (4) EXO_OPT, con las proyecciones optimistas. "
        "El motor detecta automáticamente la columna de fecha, el orden temporal (mensual, trimestral o anual), "
        "el nombre de la endógena (segunda columna) y las exógenas (desde la tercera columna en adelante). "
        "Se valida que los escenarios mantengan el mismo orden de variables que el histórico y que la proyección "
        "sea una continuación inmediata del histórico (sin huecos temporales).",
        estilo_cuerpo
    ))

    story.append(Paragraph("1.2 Bloque 1 — Exógenas: Estacionariedad y Diferenciación", estilo_subseccion))
    story.append(Paragraph(
        "Se aplica la prueba Augmented Dickey-Fuller (ADF) a cada exógena para encontrar el orden mínimo "
        "de diferenciación (d) que la vuelve estacionaria. Cada exógena puede tener su propia d, lo que "
        "permite flexibilidad en el tratamiento de variables con diferentes propiedades de persistencia. "
        "Una serie estacionaria tiene media y varianza constantes en el tiempo, propiedad fundamental para "
        "la validez de las inferencias del modelo. El resultado de este bloque es un vector dinámico de exógenas "
        "ya diferenciadas (var_exogenas_diff) y un diccionario con la d aplicada por variable (var_exogenas_d), "
        "necesario para revertir la diferenciación en el forecast.",
        estilo_cuerpo
    ))

    story.append(Paragraph("1.3 Bloque 2 — Endógena: Tipo y Diferenciación", estilo_subseccion))
    story.append(Paragraph(
        "Se evalúan dos transformaciones posibles para la variable dependiente: (a) Original (total): la serie "
        "en su escala natural, útil cuando la variable no está acotada; y (b) Logit: transformación logit(y) = ln(y/(1-y)), "
        "aplicable cuando la variable está acotada entre 0 y 1 (ej. tasas de morosidad). La transformación logit evita "
        "proyecciones fuera del rango lógico. Para cada opción se ejecuta la prueba ADF y se selecciona la que sea "
        "estacionaria con d ≤ 1, priorizando la parsimonia del modelo. El resultado es la endógena final ya "
        "diferenciada (var_endogena_diff), junto con los órdenes AR(p) y MA(q) candidatos (VALORES_P y VALORES_Q) "
        "que alimentarán el motor de combinaciones.",
        estilo_cuerpo
    ))

    story.append(Paragraph("1.4 Bloque 3 — Lags: Correlograma Cruzado", estilo_subseccion))
    story.append(Paragraph(
        "Mediante correlograma cruzado entre la endógena diferenciada y cada exógena diferenciada, se sugiere "
        "el rezago óptimo para cada variable. Este proceso respeta el signo económico esperado (positivo o negativo) "
        "configurado por el analista en la lista SIGNOS_EXOGENAS, garantizando que la relación entre la cartera y "
        "cada macro sea teóricamente coherente. El flujo contempla tres escenarios: (1) si hay lags significativos que "
        "siguen el signo esperado, se toman los N más significativos; (2) si ninguno es significativo pero siguen el signo, "
        "se toma el más cercano; (3) si no sigue el signo, se toma un solo lag, el más significativo. El resultado es un "
        "diccionario {exógena: [lags]} (lags_por_exogena) que alimenta el Bloque 5.",
        estilo_cuerpo
    ))

    story.append(Paragraph("1.5 Bloque 4 — Dummies: Variables Opcionales", estilo_subseccion))
    story.append(Paragraph(
        "Este bloque es opcional y permite generar variables dummy (vectores de 0 y 1) a partir de un rango de "
        "fechas (inicio y fin). Cada dummy vale 1 dentro del período definido y 0 fuera de él. Estas variables se "
        "construyen sobre el índice completo (histórico + proyección) para que sirvan tanto al ajuste como al forecast. "
        "Si no se definen dummies, el Bloque 5 entiende que no aplican y no las incluye en las combinaciones. "
        "Ejemplos típicos incluyen eventos excepcionales como la pandemia COVID-19 o crisis financieras.",
        estilo_cuerpo
    ))

    story.append(Paragraph("1.6 Bloque 5 — Combinaciones: Motor SARIMAX y Filtrado", estilo_subseccion))
    story.append(Paragraph(
        "El motor central genera todas las combinaciones posibles de modelos SARIMAX guiadas por los bloques "
        "anteriores. Las reglas de combinación son: subconjuntos de exógenas (máximo N por modelo, definido por "
        "MAX_EXOG_POR_MODELO); cada exógena entra con un solo lag (exclusivo, según Bloque 3); órdenes AR(p) "
        "y MA(q) candidatos (Bloque 2); y con/sin dummy (Bloque 4). Para cada especificación se ajusta un SARIMAX(p,0,q) "
        "sobre el histórico recortado a la ventana definida. Los modelos se filtran por: (a) VIF < VIF_MAX para controlar "
        "multicolinealidad; (b) signos de coeficientes coherentes con la teoría económica (estricto); (c) modelos con p=0 "
        "se marcan para revisión pero no se descartan automáticamente. El resultado es una lista de modelos aceptados "
        "(modelos_aceptados) ordenados por AIC.",
        estilo_cuerpo
    ))

    story.append(Paragraph("1.7 Bloque 6 — Forecast: Proyección de Modelos Aceptados", estilo_subseccion))
    story.append(Paragraph(
        "Se proyecta cada modelo aceptado sobre los tres escenarios (BAS/ADV/OPT). La transformación consistente "
        "de las exógenas es clave: para cada exógena se diferencia la serie [histórico + proyección] con su misma d "
        "(Bloque 1) y se aplica el mismo lag del modelo, de modo que la primera proyección enlaza con el último histórico. "
        "Luego se revierte la diferenciación de la endógena (integración) y, si aplica, la transformación logit (inversa: "
        "y = 1/(1+exp(-x))). Se aplica bloqueo de solapamiento para evitar que escenarios laterales (ADV/OPT) crucen al "
        "base, garantizando el orden económico Adverso ≤ Base ≤ Optimista (o viceversa según la variable). El resultado "
        "es un diccionario {id_modelo: DataFrame[BAS, ADV, OPT]} en la escala original de la endógena.",
        estilo_cuerpo
    ))

    story.append(Paragraph("1.8 Bloque 7 — Ordenamiento: Sensibilidad y Filtrado Final", estilo_subseccion))
    story.append(Paragraph(
        "Se descartan los modelos de baja sensibilidad: aquellos cuyas proyecciones de optimista y adverso se parecen "
        "demasiado (no diferencian los escenarios). La métrica es |media(OPT) − media(ADV)| sobre el horizonte de "
        "proyección. Se aplican cuatro filtros adicionales: (1) sensibilidad ≥ UMBRAL_SENSIBILIDAD; (2) sin outliers: "
        "todo el forecast dentro de media_hist ± K·std_hist; (3) significancia: al menos una exógena con p-value ≤ 0.05; "
        "(4) factor FWL: todos los factores (forecast / última PD observada) dentro del rango [FWL_FACTOR_MIN, FWL_FACTOR_MAX]. "
        "Los modelos que pasan todos los filtros se ordenan de mayor a menor sensibilidad y constituyen la variable final "
        "modelos_finales, insumo del Bloque 8.",
        estilo_cuerpo
    ))

    story.append(Paragraph("1.9 Bloque 8 — Exportación: Excel de Modelos Finales", estilo_subseccion))
    story.append(Paragraph(
        "Se exportan los primeros TOP_EXPORTAR modelos de modelos_finales a un archivo .xlsx, una hoja por modelo. "
        "Cada hoja replica la estructura de trazabilidad completa: endógena (histórico + forecast de los 3 escenarios); "
        "exógenas del modelo (nivel histórico común + 3 escenarios); FWL a 12 meses (factor nivel / último histórico); "
        "Factor FWL por Año (resumen de diciembre de cada año); residuos individuales y resumen de distribución; "
        "coeficientes del modelo (con p-values); y pruebas estadísticas (Ljung-Box, Jarque-Bera, ARCH). Además, se "
        "embebe metadata de trazabilidad en las propiedades del workbook (país, cartera, parámetros del motor, etc.).",
        estilo_cuerpo
    ))

    story.append(Paragraph("1.10 Bloque 9 — Reporte CSV: Factores de Impacto", estilo_subseccion))
    story.append(Paragraph(
        "Este bloque genera un reporte .csv de factores de impacto (forecast / última PD observada) para un único "
        "modelo elegido por su id, en los escenarios BAS, ADV, OPT y ORI (constante = 1). El factor se extiende de forma "
        "plana hasta una fecha de corte configurable (por defecto 2100-12-01). Previamente, una celda auxiliar (B9.1) "
        "evalúa los modelos exportados y muestra un top de candidatos cuyos factores FWL estén dentro o más cerca "
        "del rango objetivo [FWL_RANGO_MIN, FWL_RANGO_MAX], facilitando la selección del modelo a reportar.",
        estilo_cuerpo
    ))
    story.append(PageBreak())

    # =====================================================================
    # SECCIÓN 2 — CONFIGURACIÓN DEL MOTOR
    # =====================================================================
    story.append(Paragraph("2. Configuración del Motor", estilo_seccion))

    cfg_data = [
        ["Parámetro", "Valor"],
        ["País", meta.get("pais", "N/A")],
        ["Cartera", meta.get("cartera", "N/A")],
        ["Tipo de endógena", meta.get("tipo_endogena", "N/A")],
        ["Modo de endógena", meta.get("modo_endogena", "N/A")],
        ["Ventana media móvil", str(meta.get("ventana_mm", "N/A")) + " meses"],
        ["VIF máximo permitido", str(meta.get("vif_max", "N/A"))],
        ["Máx. exógenas por modelo", str(meta.get("max_exog", "N/A"))],
        ["Top modelos exportados", str(meta.get("top_exportar", "N/A"))],
        ["Órdenes AR candidatos", str(meta.get("valores_p", "N/A"))],
        ["Órdenes MA candidatos", str(meta.get("valores_q", "N/A"))],
        ["Máximo lags evaluados", str(meta.get("max_lags", "N/A"))],
        ["Umbral de sensibilidad", str(meta.get("umbral_sensibilidad", "N/A"))],
        ["Rango FWL mínimo", str(meta.get("fwl_min", "N/A"))],
        ["Rango FWL máximo", str(meta.get("fwl_max", "N/A"))],
        ["Trend del modelo", meta.get("trend_modelo", "N/A")],
    ]
    tabla_cfg = Table(cfg_data, colWidths=[2.5 * inch, 3.2 * inch])
    tabla_cfg.setStyle(tabla_estilo_base())
    story.append(tabla_cfg)
    story.append(Spacer(1, 0.2 * inch))

    story.append(Paragraph("2.1 Factor FWL (Forward-Looking)", estilo_subseccion))
    fwl_min = meta.get("fwl_min", "N/A")
    fwl_max = meta.get("fwl_max", "N/A")
    story.append(Paragraph(
        f"El Factor FWL es el multiplicador escalar que se aplica sobre la proyección base "
        f"para obtener las trayectorias adverso y optimista. El rango FWL define los límites "
        f"mínimo y máximo dentro de los cuales puede oscilar este factor, garantizando que las "
        f"proyecciones extremas no se desvíen de manera irrazonable respecto a la trayectoria base. "
        f"RANGO FWL CONFIGURADO: {fwl_min} — {fwl_max}. "
        f"Un rango más amplio implica mayor incertidumbre en los escenarios, mientras que un rango "
        f"más estrecho refleja mayor confianza en la estabilidad de la cartera.",
        estilo_cuerpo
    ))

    story.append(Paragraph("2.2 Tipo de endógena: Logit vs Total", estilo_subseccion))
    tipo_endog = meta.get("tipo_endogena", "N/A")
    if str(tipo_endog).lower() == "logit":
        texto_tipo = (
            "Logit: La variable endógena se transforma mediante la función logit antes de la estimación. "
            "Esto es útil cuando la variable está acotada entre 0 y 1 (o 0% y 100%), ya que evita que el modelo "
            "proyecte valores fuera del rango lógico. Los resultados deben retransformarse mediante la inversa "
            "del logit (exp(x)/(1+exp(x))) para su interpretación en la escala original."
        )
    else:
        texto_tipo = (
            "Total: La variable endógena se modela en su escala original (por ejemplo, "
            "porcentaje de morosidad o tasa de variación). Los coeficientes se interpretan "
            "directamente en unidades de la variable dependiente."
        )
    story.append(Paragraph(f"TIPO ENDÓGENA CONFIGURADO: {tipo_endog}. {texto_tipo}", estilo_cuerpo))

    story.append(Paragraph("2.3 Modo de endógena: Actual vs Media Móvil", estilo_subseccion))
    modo_endog = meta.get("modo_endogena", "N/A")
    ventana_mm = meta.get("ventana_mm", "N/A")
    if str(modo_endog).lower() in ["media_movil", "media móvil", "media movil"]:
        texto_modo = (
            "Media Móvil: La endógena se calcula como el promedio móvil de los últimos N meses. "
            "Este suavizado reduce el ruido de corto plazo y ayuda a identificar tendencias subyacentes, "
            "siendo preferible cuando la serie presenta picos atípicos o alta frecuencia de variaciones puntuales."
        )
    else:
        texto_modo = (
            "Actual: El modelo utiliza el valor observado de la serie en cada período como variable dependiente. "
            "Es el enfoque más directo y se recomienda cuando la serie presenta baja volatilidad de corto plazo."
        )
    story.append(Paragraph(f"MODO CONFIGURADO: {modo_endog}. VENTANA MEDIA MÓVIL: {ventana_mm} meses. {texto_modo}", estilo_cuerpo))

    story.append(Paragraph("2.4 Umbrales de sensibilidad", estilo_subseccion))
    umbral_sens = meta.get("umbral_sensibilidad", "N/A")
    story.append(Paragraph(
        f"Los umbrales de sensibilidad definen los límites de variación aceptables para los coeficientes "
        f"de las variables exógenas ante shocks en las proyecciones. Si el cambio en una macro excede el umbral "
        f"definido, el modelo alerta sobre posible inestabilidad estructural. "
        f"UMBRAL DE SENSIBILIDAD CONFIGURADO: {umbral_sens}. "
        f"Nota: Este parámetro debe verificarse contra la configuración vigente del motor de modelos.",
        estilo_cuerpo
    ))

    story.append(Paragraph("2.5 Parámetros del motor", estilo_subseccion))
    story.append(Paragraph(
        f"Los siguientes parámetros fueron utilizados en la corrida que generó este modelo: "
        f"VIF máximo permitido: {meta.get('vif_max', 'N/A')}. "
        f"Máximo de exógenas por modelo: {meta.get('max_exog', 'N/A')}. "
        f"Top de modelos a exportar: {meta.get('top_exportar', 'N/A')}. "
        f"Órdenes AR candidatos: {meta.get('valores_p', 'N/A')}. "
        f"Órdenes MA candidatos: {meta.get('valores_q', 'N/A')}. "
        f"Máximo de lags evaluados: {meta.get('max_lags', 'N/A')}. "
        f"Significancia mínima de exógenas: p ≤ 0.05. "
        f"Trend del modelo: {meta.get('trend_modelo', 'N/A')}.",
        estilo_cuerpo
    ))
    story.append(PageBreak())

    # =====================================================================
    # SECCIÓN 3 — VARIABLES UTILIZADAS
    # =====================================================================
    story.append(Paragraph("3. Variables Utilizadas", estilo_seccion))

    exo_data = [["Variable exógena", "Coeficiente", "p-value", "Significancia"]]
    exo_count = 0
    for row in coeficientes:
        if row.get("tipo") == "Exogena":
            exo_count += 1
            pval = row.get("p_value")
            try:
                sig = "Sí" if pval is not None and float(pval) < 0.05 else "No"
            except:
                sig = "No"
            coef_val = row.get("coeficiente")
            exo_data.append([
                str(row.get("variable", "N/A")),
                f"{float(coef_val):.6f}" if coef_val is not None else "N/A",
                fmt_num(pval),
                sig
            ])

    if exo_count > 0:
        for row in exo_data[1:]:
            row[0] = truncar_texto(row[0], max_len=42)
        tabla_exo = Table(exo_data, colWidths=[3.0 * inch, 1.1 * inch, 1.0 * inch, 1.0 * inch])
        tabla_exo.setStyle(tabla_estilo_base())
        story.append(tabla_exo)
    else:
        story.append(Paragraph("No se encontraron variables exógenas significativas en este modelo.", estilo_cuerpo))

    story.append(Spacer(1, 0.15 * inch))
    story.append(Paragraph("Especificación ARMA", estilo_subseccion))
    story.append(Paragraph(
        f"Términos autorregresivos (AR): {ar_count}. "
        f"Términos de media móvil (MA): {ma_count}. "
        f"Observaciones utilizadas en la estimación: {observaciones}.",
        estilo_cuerpo
    ))
    story.append(PageBreak())

    # =====================================================================
    # SECCIÓN 4 — GRÁFICA DE VARIABLES EXÓGENAS
    # =====================================================================
    story.append(Paragraph("4. Gráfica de Variables Exógenas", estilo_seccion))
    story.append(Paragraph(
        "La siguiente figura muestra la evolución histórica y proyectada de cada variable exógena "
        "incluida en el modelo, bajo los tres escenarios macroeconómicos (Base, Adverso y Optimista). "
        "Cada panel corresponde a una variable exógena diferente. La línea vertical indica el punto de corte "
        "entre el período histórico y el de proyección.",
        estilo_cuerpo
    ))
    story.append(Spacer(1, 0.15 * inch))

    if ruta_imagen_exog and os.path.exists(ruta_imagen_exog):
        try:
            from PIL import Image as PILImage
            with PILImage.open(ruta_imagen_exog) as im:
                img_w_px, img_h_px = im.size
            aspecto = img_h_px / img_w_px
            ancho_max = 6.5 * inch
            alto_max = 8.8 * inch  # margen de página disponible
            ancho = ancho_max
            alto = ancho * aspecto
            if alto > alto_max:
                alto = alto_max
                ancho = alto / aspecto
            story.append(Image(ruta_imagen_exog, width=ancho, height=alto))
        except Exception:
            story.append(Image(ruta_imagen_exog, width=6.5 * inch, height=4.5 * inch))
    else:
        story.append(Paragraph("[Gráfica no disponible]", estilo_nota))

    story.append(PageBreak())

    # =====================================================================
    # SECCIÓN 5 — DIAGNÓSTICOS E INTERPRETACIÓN
    # =====================================================================
    story.append(Paragraph("5. Diagnósticos e Interpretación", estilo_seccion))
    story.append(Paragraph(
        "A continuación se presentan los resultados de las pruebas estadísticas aplicadas a los residuos "
        "del modelo, junto con su interpretación metodológica. Cada prueba recibe un score de A a D, "
        "donde A indica cumplimiento óptimo y D indica incumplimiento significativo.",
        estilo_cuerpo
    ))
    story.append(Spacer(1, 0.1 * inch))

    # Ljung-Box
    lb_score = lb.get("score", "N/A")
    lb_p = lb.get("p_value", 0)
    lb_stat = lb.get("estadistico", 0)
    lb_interp = lb.get("interpretacion", texto_ljungbox(lb_score, lb_p if lb_p is not None else 0))
    story.append(Paragraph(f"5.1 Ljung-Box — Autocorrelación de residuos (Score: {lb_score})", estilo_subseccion))
    story.append(Paragraph(f"Estadístico: {fmt_num(lb_stat)}  |  p-valor: {fmt_num(lb_p)}", estilo_nota))
    story.append(Paragraph(lb_interp, estilo_por_score(lb_score)))
    story.append(Spacer(1, 0.1 * inch))

    # Jarque-Bera
    jb_score = jb.get("score", "N/A")
    jb_p = jb.get("p_value", 0)
    jb_stat = jb.get("estadistico", 0)
    jb_interp = jb.get("interpretacion", texto_jarquebera(jb_score, jb_p if jb_p is not None else 0))
    story.append(Paragraph(f"5.2 Jarque-Bera — Normalidad de residuos (Score: {jb_score})", estilo_subseccion))
    story.append(Paragraph(f"Estadístico: {fmt_num(jb_stat)}  |  p-valor: {fmt_num(jb_p)}", estilo_nota))
    story.append(Paragraph(jb_interp, estilo_por_score(jb_score)))
    story.append(Spacer(1, 0.1 * inch))

    # Heterocedasticidad
    ht_score = ht.get("score", "N/A")
    ht_p = ht.get("p_value", 0)
    ht_stat = ht.get("estadistico", 0)
    ht_interp = ht.get("interpretacion", texto_hetero(ht_score, ht_p if ht_p is not None else 0))
    story.append(Paragraph(f"5.3 Heterocedasticidad — ARCH/LM (Score: {ht_score})", estilo_subseccion))
    story.append(Paragraph(f"Estadístico: {fmt_num(ht_stat)}  |  p-valor: {fmt_num(ht_p)}", estilo_nota))
    story.append(Paragraph(ht_interp, estilo_por_score(ht_score)))
    story.append(PageBreak())

    # =====================================================================
    # SECCIÓN 6 — SCORE GLOBAL Y CONCLUSIÓN (solo si el modelo califica como BUENO)
    # =====================================================================
    if score_global is not None and score_global >= 7:
        story.append(Paragraph("6. Score Global y Conclusión", estilo_seccion))

        story.append(Paragraph(
            "El score global es una métrica ponderada que resume la calidad diagnóstica del modelo en una "
            "escala de 0 a 10. Se construye a partir de los scores individuales (A, B, C, D) de las tres pruebas "
            "estadísticas aplicadas a los residuos, siguiendo los siguientes pasos:",
            estilo_cuerpo
        ))
        story.append(Paragraph(
            "<b>Paso 1 — Score individual por prueba:</b> Cada prueba recibe una letra según su p-valor. "
            "Para Ljung-Box y Heterocedasticidad: A (p > 0.10), B (0.05 < p ≤ 0.10), C (0.01 < p ≤ 0.05), D (p ≤ 0.01). "
            "Para Jarque-Bera se aplican los mismos umbrales. A indica cumplimiento óptimo, D indica incumplimiento significativo.",
            estilo_cuerpo
        ))
        story.append(Paragraph(
            "<b>Paso 2 — Conversión numérica:</b> Cada letra se convierte a un valor numérico: A = 10.0, B = 7.5, C = 5.0, D = 2.5. "
            "Esta escala refleja que un score C ya representa una señal de alerta que requiere monitoreo, mientras que D indica "
            "un problema estadístico que compromete la validez del modelo.",
            estilo_cuerpo
        ))
        story.append(Paragraph(
            "<b>Paso 3 — Ponderación:</b> Los tres scores numéricos se combinan con pesos que reflejan la importancia relativa "
            "de cada diagnóstico para la robustez del modelo SARIMAX: Ljung-Box (autocorrelación de residuos) pesa 40%; "
            "Jarque-Bera (normalidad de residuos) pesa 30%; y Heterocedasticidad (varianza constante) pesa 30%. "
            "La fórmula es: <i>Score Global = 0.40 × Score_Ljung + 0.30 × Score_Jarque + 0.30 × Score_Hetero</i>.",
            estilo_cuerpo
        ))
        story.append(Paragraph(
            "<b>Paso 4 — Clasificación:</b> El resultado final se interpreta así: ≥ 7.0 = BUENO (modelo robusto para uso oficial); "
            "5.0 – 6.9 = REGULAR (utilizable con seguimiento); < 5.0 = DEFICIENTE (requiere reespecificación antes de uso oficial). "
            "Esta clasificación permite al equipo de modelos de riesgo tomar decisiones informadas sobre la selección final.",
            estilo_cuerpo
        ))
        story.append(Spacer(1, 0.1 * inch))

        resumen_data = [
            ["Prueba", "Score", "p-valor", "Estadístico"],
            ["Ljung-Box", lb_score, fmt_num(lb_p), fmt_num(lb_stat)],
            ["Jarque-Bera", jb_score, fmt_num(jb_p), fmt_num(jb_stat)],
            ["Heterocedasticidad", ht_score, fmt_num(ht_p), fmt_num(ht_stat)],
            ["Score Global Ponderado", "", "", f"{score_global:.2f} / 10"],
        ]
        tabla_resumen = Table(resumen_data, colWidths=[2.2 * inch, 1.0 * inch, 1.2 * inch, 1.3 * inch])
        tabla_resumen.setStyle(tabla_estilo_base())
        story.append(tabla_resumen)
        story.append(Spacer(1, 0.2 * inch))

        clas_norm = str(clasificacion).upper()
        if clas_norm == "BUENO":
            color_clas = VERDE_CORP
        elif clas_norm == "REGULAR":
            color_clas = AMBAR_CORP
        else:
            color_clas = ROJO_CORP

        estilo_clas = ParagraphStyle(
            "ClasificacionDoc",
            parent=estilo_seccion,
            textColor=color_clas,
            alignment=TA_CENTER,
            fontSize=18,
            spaceAfter=12
        )
        story.append(Paragraph(f"Clasificación: {clasificacion}", estilo_clas))
        story.append(Paragraph(f"Score global: {score_global:.2f} / 10", estilo_subseccion))
        story.append(Spacer(1, 0.1 * inch))
        story.append(Paragraph("Conclusión", estilo_subseccion))
        story.append(Paragraph(conclusion_global, estilo_cuerpo))

        story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph(
        "Este documento fue generado automáticamente por el dashboard SARIMAX IFRS 9. "
        "Los resultados deben ser validados por el equipo de modelos de riesgo antes de su uso oficial.",
        estilo_nota
    ))

    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes
def generar_excel_metodologico(doc_data, ruta_imagen):
    wb = openpyxl.Workbook()
    ws_meta = wb.active
    ws_meta.title = "Metadatos"
    ws_meta.append(["Campo", "Valor"])
    meta_rows = [
        ("Modelo final", doc_data.get("nombre_modelo", "N/A")),
        ("Fecha generacion", doc_data.get("fecha_generacion", "N/A")),
        ("Pais", doc_data.get("meta_kpis", {}).get("pais", "N/A")),
        ("Cartera", doc_data.get("meta_kpis", {}).get("cartera", "N/A")),
        ("Tipo endogena", doc_data.get("meta_kpis", {}).get("tipo_endogena", "N/A")),
        ("Modo endogena", doc_data.get("meta_kpis", {}).get("modo_endogena", "N/A")),
        ("Ventana MM", doc_data.get("meta_kpis", {}).get("ventana_mm", "N/A")),
        ("VIF max", doc_data.get("vif_max", "N/A")),
        ("Rango FWL", f"{doc_data.get('fwl_min', 'N/A')} - {doc_data.get('fwl_max', 'N/A')}"),
        ("Max exogenas", doc_data.get("max_exog", "N/A")),
        ("Top exportar", doc_data.get("top_exportar", "N/A")),
        ("Umbral sensibilidad", doc_data.get("umbral_sensibilidad", "N/A")),
        ("AR / MA", f"{doc_data.get('ar_count', 0)} / {doc_data.get('ma_count', 0)}"),
        ("Score global", f"{doc_data.get('score_global_num')}/10" if doc_data.get("score_global_num") is not None else "N/A"),
        ("Clasificacion score", doc_data.get("score_global_clase", "N/A")),
        ("Conclusion", doc_data.get("score_global_conclusion", "")),
    ]
    for row in meta_rows:
        ws_meta.append(list(row))

    ws_exo = wb.create_sheet("Exógenas")
    ws_exo.append(["Exogena", "P_value", "Estado"])
    for row in doc_data.get("exogenas", []):
        ws_exo.append([row.get("exogena"), row.get("p_value"), row.get("estado")])

    ws_diag = wb.create_sheet("Diagnósticos")
    ws_diag.append(["Prueba", "Estadistico", "P_value", "Score", "Interpretacion"])
    for row in doc_data.get("diagnosticos", []):
        ws_diag.append([
            row.get("prueba"), row.get("estadistico"), row.get("p_value"),
            row.get("score"), row.get("interpretacion")
        ])

    ws_coef = wb.create_sheet("Coeficientes")
    ws_coef.append(["Variable", "Coeficiente", "P_value", "Tipo"])
    for row in doc_data.get("coeficientes", []):
        ws_coef.append([row.get("variable"), row.get("coeficiente"), row.get("p_value"), row.get("tipo")])

    ws_graph = wb.create_sheet("Gráfica")
    ws_graph["A1"] = "Grafica de macros (solo exogenas)"
    if ruta_imagen and os.path.exists(ruta_imagen):
        img = XLImage(ruta_imagen)
        img.width = 920
        img.height = 420
        ws_graph.add_image(img, "A3")

    for ws in [ws_meta, ws_exo, ws_diag, ws_coef]:
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 60

    buff = BytesIO()
    wb.save(buff)
    buff.seek(0)
    return buff.getvalue()

def generar_y_guardar_documentos(nombre_modelo):
    """Genera PDF + Excel, guarda en session_state, retorna True/False."""
    try:
        doc_data = recolectar_datos_documento(
            nombre_modelo,
            st.session_state.modelos_data,
            st.session_state.meta_contexto
        )
    except Exception as e:
        st.error(f"Error recolectando datos: {e}")
        return False

    ruta_tmp_exog = None
    ruta_tmp_fwl = None
    ruta_tmp_hist = None
    try:
        tmp_exog = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        ruta_tmp_exog = tmp_exog.name
        tmp_exog.close()
        graficar_macros_estatico(doc_data, ruta_tmp_exog)
    except Exception as e:
        st.warning(f"Gráfica de exógenas no generada: {e}")
        ruta_tmp_exog = None

    try:
        tmp_fwl = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        ruta_tmp_fwl = tmp_fwl.name
        tmp_fwl.close()
        graficar_fwl_estatico(doc_data, ruta_tmp_fwl)
    except Exception as e:
        st.warning(f"Gráfica de FWL no generada: {e}")
        ruta_tmp_fwl = None

    try:
        tmp_hist = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
        ruta_tmp_hist = tmp_hist.name
        tmp_hist.close()
        graficar_historico_estatico(doc_data, ruta_tmp_hist)
    except Exception as e:
        st.warning(f"Gráfica histórica no generada: {e}")
        ruta_tmp_hist = None

    try:
        pdf_bytes = generar_pdf_metodologico(doc_data, ruta_tmp_exog, ruta_tmp_fwl, ruta_tmp_hist)
    except Exception as e:
        st.error(f"Error generando PDF: {e}")
        st.exception(e)
        if ruta_tmp_exog and os.path.exists(ruta_tmp_exog):
            os.remove(ruta_tmp_exog)
        if ruta_tmp_fwl and os.path.exists(ruta_tmp_fwl):
            os.remove(ruta_tmp_fwl)
        return False

    try:
        excel_bytes = generar_excel_metodologico(doc_data, ruta_tmp_exog)
    except Exception as e:
        st.error(f"Error generando Excel: {e}")
        if ruta_tmp_exog and os.path.exists(ruta_tmp_exog):
            os.remove(ruta_tmp_exog)
        if ruta_tmp_fwl and os.path.exists(ruta_tmp_fwl):
            os.remove(ruta_tmp_fwl)
        return False

    if ruta_tmp_exog and os.path.exists(ruta_tmp_exog):
        os.remove(ruta_tmp_exog)
    if ruta_tmp_fwl and os.path.exists(ruta_tmp_fwl):
        os.remove(ruta_tmp_fwl)
    if ruta_tmp_hist and os.path.exists(ruta_tmp_hist):
        os.remove(ruta_tmp_hist)

    nombre_archivo = _normalizar_nombre_archivo(nombre_modelo)
    st.session_state.modelo_final = nombre_modelo
    st.session_state.documento_metodologico_pdf = pdf_bytes
    st.session_state.documento_metodologico_excel = excel_bytes
    st.session_state.documento_metodologico_pdf_nombre = f"Documento_Metodologico_{nombre_archivo}.pdf"
    st.session_state.documento_metodologico_excel_nombre = f"Documento_Metodologico_{nombre_archivo}.xlsx"
    st.session_state.documento_metodologico_data = doc_data
    return True

def aplicar_tema_plotly(fig):
    fig.update_layout(
        font=dict(family="Inter, Arial, sans-serif", size=12, color=TEXT),
        plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
        legend=dict(bgcolor="rgba(0,0,0,0)"),
        margin=dict(t=40, l=10, r=10, b=10),
    )
    fig.update_xaxes(showgrid=True, gridcolor=BORDER, zeroline=False, linecolor=BORDER)
    fig.update_yaxes(showgrid=True, gridcolor=BORDER, zeroline=False, linecolor=BORDER)
    if fig.layout.title and fig.layout.title.text:
        fig.update_layout(title=dict(font=dict(size=14, color=NAVY)))
    return fig

def fig_predicciones(df_end, endogena_cols, exog_df, exog_sel, modelo_nombre):
    fig = go.Figure()
    fecha_col = 'fecha' if 'fecha' in df_end.columns else df_end.columns[0]
    for col in endogena_cols:
        if col not in df_end.columns:
            continue
        col_str = str(col).upper()
        if col_str == 'BASE':
            fig.add_trace(go.Scatter(x=df_end[fecha_col], y=df_end[col], mode='lines', name='Base', line=dict(color=BLUE, width=2)))
        elif col_str in ['ADVERSO', 'ADVERSA']:
            fig.add_trace(go.Scatter(x=df_end[fecha_col], y=df_end[col], mode='lines', name='Adverso', line=dict(color=RED, width=2, dash='dash')))
        elif col_str == 'OPTIMISTA':
            fig.add_trace(go.Scatter(x=df_end[fecha_col], y=df_end[col], mode='lines', name='Optimista', line=dict(color=GREEN, width=2, dash='dot')))
    if exog_df is not None and exog_sel:
        for ex in exog_sel:
            for suffix in ['_BASE', '_ADVERSO', '_OPTIMISTA']:
                col_name = ex + suffix
                if col_name in exog_df.columns:
                    x_vals = exog_df[fecha_col] if fecha_col in exog_df.columns else exog_df.index
                    fig.add_trace(go.Scatter(x=x_vals, y=exog_df[col_name], mode='lines', name=f'{ex}{suffix}', line=dict(width=1.2), yaxis='y2'))
        fig.update_layout(yaxis2=dict(title='Exogenas', overlaying='y', side='right'))
    fig.update_layout(title=f"Predicciones - {modelo_nombre}", xaxis_title="Fecha", yaxis_title="Valor",
                      legend=dict(orientation='v', yanchor='top', y=1, xanchor='left', x=1.05), hovermode='x unified')
    return aplicar_tema_plotly(fig)

def fig_fwl_12m(df_fwl):
    fig = go.Figure()
    fecha_col = None
    for c in df_fwl.columns:
        if 'fecha' in str(c).lower():
            fecha_col = c
            break
    if fecha_col is None:
        fecha_col = df_fwl.columns[0]
    for col in df_fwl.columns:
        col_str = str(col).upper()
        if 'FWL_BASE' in col_str:
            fig.add_trace(go.Scatter(x=df_fwl[fecha_col], y=df_fwl[col], mode='lines', name='Base', line=dict(color=BLUE, width=2)))
        elif 'FWL_ADVERSO' in col_str or 'FWL_ADVERSA' in col_str:
            fig.add_trace(go.Scatter(x=df_fwl[fecha_col], y=df_fwl[col], mode='lines', name='Adverso', line=dict(color=RED, width=2, dash='dash')))
        elif 'FWL_OPTIMISTA' in col_str:
            fig.add_trace(go.Scatter(x=df_fwl[fecha_col], y=df_fwl[col], mode='lines', name='Optimista', line=dict(color=GREEN, width=2, dash='dot')))
    fig.update_layout(title="Factor FWL a 12 Meses", xaxis_title="Fecha", yaxis_title="FWL",
                      legend=dict(orientation='v', yanchor='top', y=1, xanchor='left', x=1.05), hovermode='x unified')
    return aplicar_tema_plotly(fig)

def fig_fwl_ponderado(df_pond):
    fecha_col = None
    for c in df_pond.columns:
        if 'fecha' in str(c).lower():
            fecha_col = c
            break
    if fecha_col is None:
        fecha_col = df_pond.columns[0]
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=df_pond[fecha_col], y=df_pond['FWL_Ponderado'], mode='lines', name='FWL Ponderado',
                              fill='tozeroy', line=dict(color=BLUE, width=2), fillcolor='rgba(30,90,168,0.10)'))
    fig.update_layout(title="Factor FWL Ponderado", xaxis_title="Fecha", yaxis_title="FWL Ponderado")
    return aplicar_tema_plotly(fig)

def fig_histograma_residuos(vals, media, std):
    fig = go.Figure()
    fig.add_trace(go.Histogram(x=vals, nbinsx=20, marker_color=BLUE, opacity=0.7, name='Residuos'))
    x_norm, y_norm = generar_campana_normal(vals, media, std)
    if len(x_norm) > 0:
        bin_width = (vals.max() - vals.min()) / 20 if vals.max() != vals.min() else 1
        y_norm_scaled = y_norm * len(vals) * bin_width
        fig.add_trace(go.Scatter(x=x_norm, y=y_norm_scaled, mode='lines', name='Normal teorica', line=dict(color=RED, width=2)))
    fig.update_layout(xaxis_title="Residuos", yaxis_title="Frecuencia")
    return aplicar_tema_plotly(fig)

def fig_barras_coeficientes(df_coef):
    df = df_coef.copy()
    df['Coeficiente'] = pd.to_numeric(df['Coeficiente'], errors='coerce')
    df = df.dropna(subset=['Coeficiente'])
    df['abs'] = df['Coeficiente'].abs()
    df = df.sort_values('abs', ascending=True)
    colors = [GREEN if c >= 0 else RED for c in df['Coeficiente']]
    fig = go.Figure()
    fig.add_trace(go.Bar(y=df['Variable'], x=df['Coeficiente'], orientation='h', marker_color=colors,
                          text=df['Coeficiente'].round(4), textposition='outside'))
    fig.update_layout(title="Coeficientes del Modelo", xaxis_title="Valor", yaxis_title="Variable", showlegend=False)
    return aplicar_tema_plotly(fig)

# =============================================================================
# DIAGNOSTICOS
# =============================================================================
def limpiar_nombre_prueba(nombre):
    nl = str(nombre).lower()
    if 'arch' in nl: return "Heterocedasticidad"
    if 'ljung' in nl or 'box' in nl: return "Ljung-Box"
    if 'jarque' in nl or 'bera' in nl: return "Jarque-Bera"
    return str(nombre)

def evaluar_prueba(prueba, p_val):
    score, (bg, fg) = calcular_score(p_val, prueba)
    if score == 'N/A':
        return "N/A", ESTADO_NEUTRAL, score, bg, fg
    if score == 'A':
        estado = "CUMPLE"
    elif score == 'B':
        estado = "CUMPLE"
    elif score == 'C':
        estado = "REVISAR"
    else:
        estado = "NO CUMPLE"
    return estado, (bg, fg, estado), score, bg, fg

def render_diagnosticos_corporativo(pruebas_df, mostrar_detalle_tecnico=True):
    if pruebas_df is None or pruebas_df.empty:
        st.info("No hay datos de pruebas estadisticas.")
        return
    df = pruebas_df.copy()
    df['Prueba'] = df['Prueba'].apply(limpiar_nombre_prueba)
    st.markdown(section_title("Resumen de Diagnosticos"), unsafe_allow_html=True)
    filas = []
    for _, row in df.iterrows():
        prueba = row['Prueba']
        p_val = row['P_value']
        estado, _, score, bg, fg = evaluar_prueba(prueba, p_val)
        filas.append({
            'Diagnostico': prueba,
            'Score': score,
            'ScoreBg': bg,
            'ScoreFg': fg,
            'P-valor': p_val,
            'Estadistico': row.get('Estadistico', '-'),
        })
    cols = st.columns(len(filas))
    for i, f in enumerate(filas):
        with cols[i]:
            st.markdown(f"""
            <div style="background:{f['ScoreBg']};border:1px solid {BORDER};border-radius:6px;padding:16px;text-align:center;">
                <p style="font-size:10px;color:{MUTED};margin:0 0 6px;text-transform:uppercase;letter-spacing:0.5px;font-weight:600;">{f['Diagnostico']}</p>
                <p style="font-size:28px;font-weight:700;color:{f['ScoreFg']};margin:0;">{f['Score']}</p>
                <p style="font-size:11px;color:{MUTED};margin:4px 0 0;">p = {fmt_pvalor(f['P-valor'])}</p>
            </div>
            """, unsafe_allow_html=True)
            with st.expander("Ver interpretacion"):
                st.markdown(
                    f"<p style='font-size:12px;color:{TEXT};margin:0;'>{interpretar_prueba(f['Diagnostico'], f['P-valor'], f['Score'])}</p>",
                    unsafe_allow_html=True
                )
    if not mostrar_detalle_tecnico:
        return
    st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)
    st.markdown(section_title("Detalle Tecnico"), unsafe_allow_html=True)
    df_tec = pd.DataFrame(filas)[['Diagnostico', 'Estadistico', 'P-valor', 'Score']]
    df_tec['P-valor'] = df_tec['P-valor'].apply(fmt_pvalor)
    df_tec['Estadistico'] = df_tec['Estadistico'].apply(fmt_pvalor)
    def color_score(val):
        if val == "A": return f"color: {GREEN}; font-weight: 700;"
        elif val == "B": return f"color: {BLUE}; font-weight: 700;"
        elif val == "C": return f"color: #B8860B; font-weight: 700;"
        elif val == "D": return f"color: {RED}; font-weight: 700;"
        return ""
    styler = df_tec.style
    if hasattr(styler, "map"):
        styler = styler.map(color_score, subset=['Score'])
    else:
        styler = styler.applymap(color_score, subset=['Score'])
    st.dataframe(styler, use_container_width=True, hide_index=True)

def fmt_pvalor(v):
    try:
        if pd.isna(v): return "-"
        vf = float(v)
        return f"{vf:.4f}" if vf >= 0.001 else f"{vf:.2e}"
    except: return str(v)

def render_metricas_diagnostico(pruebas_df):
    if pruebas_df is None or pruebas_df.empty:
        return
    scores = obtener_scores_modelo(pruebas_df)
    score_global, _ = calcular_score_global(pruebas_df)
    # NOTA: estilo_score_global() devuelve 3 valores (etiqueta, color, fondo),
    # a diferencia de clasificar_score_global() que devuelve solo 2
    # (etiqueta, conclusion). Usar la funcion correcta evita el
    # "ValueError: not enough values to unpack".
    etiqueta_g, color_g, _ = estilo_score_global(score_global)
    c0, c1, c2, c3 = st.columns(4)
    with c0:
        valor_g = f"{score_global:.1f}/10" if score_global is not None else "N/A"
        st.markdown(card_kpi("Score Global", valor_g, etiqueta_g, accent=color_g), unsafe_allow_html=True)
    with c1:
        score, (bg, fg) = scores.get('ljung_box', ('N/A', ESTADO_NEUTRAL))
        st.markdown(card_metric("Ljung-Box (Score)", score, fg if score != 'N/A' else TEXT), unsafe_allow_html=True)
    with c2:
        score, (bg, fg) = scores.get('jarque_bera', ('N/A', ESTADO_NEUTRAL))
        st.markdown(card_metric("Jarque-Bera (Score)", score, fg if score != 'N/A' else TEXT), unsafe_allow_html=True)
    with c3:
        score, (bg, fg) = scores.get('heterocedasticidad', ('N/A', ESTADO_NEUTRAL))
        st.markdown(card_metric("Heterocedast.", score, fg if score != 'N/A' else TEXT), unsafe_allow_html=True)

def es_favorito(nombre):
    return nombre in st.session_state.get("favoritos", set())

def alternar_favorito(nombre):
    if nombre in st.session_state.favoritos:
        st.session_state.favoritos.discard(nombre)
    else:
        st.session_state.favoritos.add(nombre)

def boton_favorito(nombre, key_suffix=""):
    activo = es_favorito(nombre)
    label = "Favorito" if activo else "Marcar favorito"
    if st.button(label, key=f"fav_{key_suffix}_{nombre}", use_container_width=True):
        alternar_favorito(nombre)
        st.rerun()

def construir_opciones_modelos():
    criterio = st.session_state.get("criterio_ordenamiento", "Pruebas aprobadas ↓")
    filtro_ljung = st.session_state.get("filtro_ljung", "Todos")
    filtro_jarque = st.session_state.get("filtro_jarque", "Todos")
    filtro_hetero = st.session_state.get("filtro_hetero", "Todos")
    filtro_favoritos = st.session_state.get("filtro_favoritos", False)
    modelos_con_pruebas = []
    for nombre, datos in st.session_state.modelos_data.items():
        pruebas = datos.get('pruebas')
        apr, tot = contar_pruebas_aprobadas(pruebas)
        scores = obtener_scores_modelo(pruebas)
        pasa_filtro = True
        if filtro_ljung != "Todos":
            score_ljung, _ = scores.get('ljung_box', ('N/A', None))
            if filtro_ljung == "A o B (Cumple)" and score_ljung not in ['A', 'B']:
                pasa_filtro = False
            elif filtro_ljung == "A, B o C" and score_ljung not in ['A', 'B', 'C']:
                pasa_filtro = False
            elif filtro_ljung == "Solo A" and score_ljung != 'A':
                pasa_filtro = False
        if filtro_jarque != "Todos":
            score_jarque, _ = scores.get('jarque_bera', ('N/A', None))
            if filtro_jarque == "A o B (Cumple)" and score_jarque not in ['A', 'B']:
                pasa_filtro = False
            elif filtro_jarque == "A, B o C" and score_jarque not in ['A', 'B', 'C']:
                pasa_filtro = False
            elif filtro_jarque == "Solo A" and score_jarque != 'A':
                pasa_filtro = False
        if filtro_hetero != "Todos":
            score_hetero, _ = scores.get('heterocedasticidad', ('N/A', None))
            if filtro_hetero == "A o B (Cumple)" and score_hetero not in ['A', 'B']:
                pasa_filtro = False
            elif filtro_hetero == "A, B o C" and score_hetero not in ['A', 'B', 'C']:
                pasa_filtro = False
            elif filtro_hetero == "Solo A" and score_hetero != 'A':
                pasa_filtro = False
        if filtro_favoritos and nombre not in st.session_state.get("favoritos", set()):
            pasa_filtro = False
        if pasa_filtro:
            score_global, _ = calcular_score_global(pruebas)
            modelos_con_pruebas.append((nombre, apr, tot, scores, score_global))
    if criterio == "Nombre (A-Z)":
        modelos_con_pruebas.sort(key=lambda x: x[0])
    elif criterio == "Pruebas aprobadas ↓":
        modelos_con_pruebas.sort(key=lambda x: (-x[1], x[0]))
    elif criterio == "Pruebas aprobadas ↑":
        modelos_con_pruebas.sort(key=lambda x: (x[1], x[0]))
    elif criterio == "Score global ↓":
        modelos_con_pruebas.sort(key=lambda x: (-(x[4] if x[4] is not None else -1), x[0]))
    elif criterio == "Score global ↑":
        modelos_con_pruebas.sort(key=lambda x: ((x[4] if x[4] is not None else 999), x[0]))
    else:
        modelos_con_pruebas.sort(key=lambda x: (x[1], x[0]))
    modelos_list = [m[0] for m in modelos_con_pruebas]
    pruebas_dict = {m[0]: (m[1], m[2]) for m in modelos_con_pruebas}
    scores_dict = {m[0]: m[3] for m in modelos_con_pruebas}
    global_dict = {m[0]: m[4] for m in modelos_con_pruebas}
    return modelos_list, pruebas_dict, scores_dict, global_dict

def label_modelo(nombre, pruebas_dict, scores_dict=None, global_dict=None):
    apr, tot = pruebas_dict.get(nombre, (0, 3))
    prefijo = "[F] " if es_favorito(nombre) else ""
    label = f"{prefijo}{nombre}  ({apr}/{tot})"
    if scores_dict and nombre in scores_dict:
        scores = scores_dict[nombre]
        mini_scores = []
        for key in ['ljung_box', 'jarque_bera', 'heterocedasticidad']:
            score, _ = scores.get(key, ('N/A', None))
            if score != 'N/A':
                mini_scores.append(score)
        if mini_scores:
            label += f"  [{'|'.join(mini_scores)}]"
    if global_dict and nombre in global_dict and global_dict[nombre] is not None:
        label += f"  - {global_dict[nombre]:.1f}/10"
    return label

def render_columna_comparacion(nombre):
    datos = st.session_state.modelos_data.get(nombre, {})
    pruebas = datos.get('pruebas')
    score_global, _ = calcular_score_global(pruebas)
    scores = obtener_scores_modelo(pruebas)
    st.markdown(f"<p style='font-weight:700;color:{NAVY};font-size:13px;margin:0 0 6px;'>{nombre}</p>", unsafe_allow_html=True)
    st.markdown(score_global_badge(score_global), unsafe_allow_html=True)
    st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)
    sc1, sc2, sc3 = st.columns(3)
    with sc1:
        s, _ = scores.get('ljung_box', ('N/A', ESTADO_NEUTRAL))
        st.markdown(card_metric("Ljung-Box", s), unsafe_allow_html=True)
    with sc2:
        s, _ = scores.get('jarque_bera', ('N/A', ESTADO_NEUTRAL))
        st.markdown(card_metric("Jarque-Bera", s), unsafe_allow_html=True)
    with sc3:
        s, _ = scores.get('heterocedasticidad', ('N/A', ESTADO_NEUTRAL))
        st.markdown(card_metric("Heterocedast.", s), unsafe_allow_html=True)
    st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)
    df_fwl_comp = datos.get('fwl_12m')
    if df_fwl_comp is not None and not df_fwl_comp.empty:
        fig = fig_fwl_12m(df_fwl_comp)
        fig.update_layout(height=280, showlegend=False, title="Factor FWL a 12 Meses")
        st.plotly_chart(fig, use_container_width=True, key=f"cmp_fig_{nombre}")
    else:
        st.caption("Sin datos de FWL a 12 meses.")
    coefs = datos.get('coeficientes')
    ar_count, ma_count = contar_ar_ma(coefs) if coefs is not None else (0, 0)
    exogenas = datos.get('exogenas_nombres', [])
    sigs = obtener_significancia_exogenas(coefs, exogenas)
    sig_count = sum(1 for _, _, s in sigs if s == "Significativa")
    mc1, mc2 = st.columns(2)
    with mc1:
        st.markdown(card_metric("AR / MA", f"{ar_count} / {ma_count}", BLUE), unsafe_allow_html=True)
    with mc2:
        st.markdown(card_metric("Exog. significativas", f"{sig_count}/{len(exogenas)}", GREEN), unsafe_allow_html=True)

def fig_distribucion_scores(modelos_data):
    nombre_map = {'ljung_box': 'Ljung-Box', 'jarque_bera': 'Jarque-Bera', 'heterocedasticidad': 'Heterocedasticidad'}
    conteo = {label: {'A': 0, 'B': 0, 'C': 0, 'D': 0} for label in nombre_map.values()}
    for datos in modelos_data.values():
        scores = obtener_scores_modelo(datos.get('pruebas'))
        for clave, label in nombre_map.items():
            letra, _ = scores.get(clave, ('N/A', None))
            if letra in conteo[label]:
                conteo[label][letra] += 1
    fig = go.Figure()
    for letra in ['A', 'B', 'C', 'D']:
        color = SCORE_COLORS[letra][1]
        fig.add_trace(go.Bar(
            name=letra, x=list(nombre_map.values()),
            y=[conteo[label][letra] for label in nombre_map.values()],
            marker_color=color
        ))
    fig.update_layout(barmode='stack', title="Distribucion de Scores por Prueba", yaxis_title="Cantidad de modelos", legend_title_text="Score")
    return aplicar_tema_plotly(fig)

def render_resumen_ejecutivo():
    modelos_data = st.session_state.modelos_data
    filas = []
    for nombre, datos in modelos_data.items():
        pruebas = datos.get('pruebas')
        score_global, _ = calcular_score_global(pruebas)
        coefs = datos.get('coeficientes')
        ar_count, ma_count = contar_ar_ma(coefs) if coefs is not None else (0, 0)
        filas.append({'Modelo': nombre, 'Score': score_global, 'AR': ar_count, 'MA': ma_count})
    df_res = pd.DataFrame(filas)
    total = len(df_res)
    con_score = df_res.dropna(subset=['Score'])
    buenos = int((con_score['Score'] >= 7).sum()) if not con_score.empty else 0
    regulares = int(((con_score['Score'] >= 5) & (con_score['Score'] < 7)).sum()) if not con_score.empty else 0
    deficientes = int((con_score['Score'] < 5).sum()) if not con_score.empty else 0
    st.markdown(f"""
    <div style="margin-bottom:8px;">
        <p style="font-size:20px;font-weight:700;color:{NAVY};margin:0;">Resumen de la corrida</p>
        <p style="font-size:12px;color:{MUTED};margin:4px 0 0;">Vista general de los {total} modelos cargados en el archivo.</p>
    </div>
    <div style="height:1px;background:{BORDER};margin:12px 0 20px;"></div>
    """, unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(card_kpi("Total de modelos", str(total)), unsafe_allow_html=True)
    with c2:
        st.markdown(card_kpi("Buenos (score >= 7)", str(buenos), accent=GREEN), unsafe_allow_html=True)
    with c3:
        st.markdown(card_kpi("Regulares (5 - 7)", str(regulares), accent="#B8860B"), unsafe_allow_html=True)
    with c4:
        st.markdown(card_kpi("Deficientes (< 5)", str(deficientes), accent=RED), unsafe_allow_html=True)
    st.markdown(divider(), unsafe_allow_html=True)
    st.plotly_chart(fig_distribucion_scores(modelos_data), use_container_width=True)
    st.markdown(divider(), unsafe_allow_html=True)
    tcol1, tcol2 = st.columns(2)
    with tcol1:
        st.markdown(section_title("Top 5 - mejor score global"), unsafe_allow_html=True)
        top5 = con_score.sort_values('Score', ascending=False).head(5)[['Modelo', 'Score']]
        st.dataframe(top5, use_container_width=True, hide_index=True)
    with tcol2:
        st.markdown(section_title("Bottom 5 - peor score global"), unsafe_allow_html=True)
        bottom5 = con_score.sort_values('Score', ascending=True).head(5)[['Modelo', 'Score']]
        st.dataframe(bottom5, use_container_width=True, hide_index=True)
    st.markdown(divider(), unsafe_allow_html=True)
    st.markdown(section_title("Estructura promedio de los modelos"), unsafe_allow_html=True)
    ac1, ac2 = st.columns(2)
    with ac1:
        st.markdown(card_metric("Promedio terminos AR", f"{df_res['AR'].mean():.2f}" if total else "0", BLUE), unsafe_allow_html=True)
    with ac2:
        st.markdown(card_metric("Promedio terminos MA", f"{df_res['MA'].mean():.2f}" if total else "0", BLUE), unsafe_allow_html=True)
    st.markdown("<div style='height:24px;'></div>", unsafe_allow_html=True)
    bexp1, bexp2 = st.columns(2)
    with bexp1:
        if st.button("Explorar modelos", key="btn_explorar_modelos", use_container_width=True):
            st.session_state.vista_resumen = False
            st.rerun()
    with bexp2:
        n_fav = len([m for m in st.session_state.get("favoritos", set()) if m in st.session_state.modelos_data])
        if st.button(f"Ver favoritos ({n_fav})", key="btn_ver_favoritos_resumen", use_container_width=True):
            st.session_state.vista_resumen = False
            st.session_state.vista_favoritos = True
            st.rerun()

def render_vista_favoritos():
    favoritos = st.session_state.get("favoritos", set())
    modelos_data = st.session_state.modelos_data
    favoritos_validos = [m for m in favoritos if m in modelos_data]
    st.markdown(f"""
    <div style="margin-bottom:8px;">
        <p style="font-size:20px;font-weight:700;color:{NAVY};margin:0;">Modelos favoritos</p>
        <p style="font-size:12px;color:{MUTED};margin:4px 0 0;">{len(favoritos_validos)} modelo(s) marcados como favoritos.</p>
    </div>
    <div style="height:1px;background:{BORDER};margin:12px 0 20px;"></div>
    """, unsafe_allow_html=True)
    bcol1, bcol2 = st.columns([1, 5])
    with bcol1:
        if st.button("Volver", key="btn_volver_de_favoritos", use_container_width=True):
            st.session_state.vista_favoritos = False
            st.session_state.vista_resumen = True
            st.rerun()
    if not favoritos_validos:
        st.info("Aun no ha marcado ningun modelo como favorito. Abra un modelo y presione 'Marcar favorito' en la barra lateral, o use el boton en cada tarjeta.")
        return
    filas = []
    for nombre in favoritos_validos:
        datos = modelos_data.get(nombre, {})
        pruebas = datos.get('pruebas')
        score_global, _ = calcular_score_global(pruebas)
        filas.append((nombre, score_global))
    filas.sort(key=lambda x: (-(x[1] if x[1] is not None else -1), x[0]))
    n_cols = 3
    for i in range(0, len(filas), n_cols):
        fila_cols = st.columns(n_cols)
        for j, (nombre, score_global) in enumerate(filas[i:i + n_cols]):
            with fila_cols[j]:
                etiqueta_g, color_g, bg_g = estilo_score_global(score_global)
                datos = modelos_data.get(nombre, {})
                coefs = datos.get('coeficientes')
                ar_count, ma_count = contar_ar_ma(coefs) if coefs is not None else (0, 0)
                obs = datos.get('observaciones', 0)
                st.markdown(f"""
                <div style="background:{WHITE};border:1px solid {BORDER};border-radius:8px;padding:16px;margin-bottom:10px;">
                    <p style="font-size:13px;font-weight:700;color:{NAVY};margin:0 0 8px;">[F] {nombre}</p>
                    {score_global_badge(score_global)}
                    <p style="font-size:11px;color:{MUTED};margin:10px 0 0;">{obs} observaciones - AR {ar_count} / MA {ma_count}</p>
                </div>
                """, unsafe_allow_html=True)
                bc1, bc2 = st.columns(2)
                with bc1:
                    if st.button("Abrir", key=f"abrir_fav_{nombre}", use_container_width=True):
                        st.session_state.modelo_seleccionado = nombre
                        st.session_state.vista_favoritos = False
                        st.session_state.vista_resumen = False
                        st.rerun()
                with bc2:
                    if st.button("Quitar", key=f"quitar_fav_{nombre}", use_container_width=True):
                        alternar_favorito(nombre)
                        st.rerun()

def render_seccion_coeficientes(datos, key_prefix="diag"):
    st.markdown(section_title("Coeficientes del modelo"), unsafe_allow_html=True)
    coefs = datos.get('coeficientes')
    if coefs is not None and not coefs.empty:
        df_coef = coefs.copy()
        df_coef['Tipo'] = df_coef['Variable'].apply(clasificar_variable)
        if 'P_value' in df_coef.columns:
            def fmt_pval(x):
                if pd.isna(x): return "N/A"
                try:
                    xv = float(x)
                    return f"{xv:.4e}" if xv < 0.001 else f"{xv:.4f}"
                except: return str(x)
            df_coef['P-valor'] = df_coef['P_value'].apply(fmt_pval)
        df_display = df_coef[['Tipo', 'Variable', 'Coeficiente', 'P-valor']]
        st.plotly_chart(fig_barras_coeficientes(df_coef), use_container_width=True, key=f"{key_prefix}_coef_bar")
        def color_pval(v):
            try: return f"color: {GREEN}; font-weight: 700;" if float(v) < 0.05 else f"color: {RED};"
            except: return ""
        styler = df_display.style
        if hasattr(styler, "map"):
            styler = styler.map(color_pval, subset=['P-valor'])
        else:
            styler = styler.applymap(color_pval, subset=['P-valor'])
        st.dataframe(styler, use_container_width=True, hide_index=True, key=f"{key_prefix}_coef_tabla")
        ar_count, ma_count = contar_ar_ma(coefs)
        st.markdown(divider(), unsafe_allow_html=True)
        st.markdown(section_title("Estructura del modelo"), unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            st.markdown(card_metric("Terminos AR", str(ar_count), BLUE), unsafe_allow_html=True)
        with c2:
            st.markdown(card_metric("Terminos MA", str(ma_count), BLUE), unsafe_allow_html=True)
    else:
        st.info("No hay datos de coeficientes.")

# =============================================================================
# GENERADOR DE NOTEBOOKS
# =============================================================================

def render_generador():
    st.markdown(f"<p style='font-size:20px;font-weight:700;color:{NAVY};margin:0 0 8px;'>Generador de Notebooks SARIMAX</p>", unsafe_allow_html=True)
    st.markdown(f"<p style='font-size:12px;color:{MUTED};margin:0 0 20px;'>Configura parametros y descarga los notebooks listos para ejecutar.</p>", unsafe_allow_html=True)

    # Directorio de templates
    TEMPLATES_DIR = Path(__file__).parent
    TEMPLATE_GENERADOR = TEMPLATES_DIR / "Generacion_Variacion__2_.ipynb"
    TEMPLATE_MOTOR = TEMPLATES_DIR / "Motor_Sarimax_Vivi_CO__1_.ipynb"

    templates_ok = TEMPLATE_GENERADOR.exists() and TEMPLATE_MOTOR.exists()
    if not templates_ok:
        st.error("Falta un template. Verifica que estos archivos esten en el mismo directorio:")
        st.code(f"{TEMPLATE_GENERADOR.name}\n{TEMPLATE_MOTOR.name}")
        return

    col1, col2 = st.columns(2)

    with col1:
        st.markdown(section_title("Ubicacion"), unsafe_allow_html=True)
        pais_display = st.selectbox(
            "Pais",
            options=list(PAIS_MAP_GEN.keys()),
            help="Selecciona el pais del modelo"
        )
        pais = PAIS_MAP_GEN[pais_display]

        cartera_display = st.selectbox(
            "Cartera (selecciona o escribe manualmente abajo)",
            options=list(CARTERA_MAP_GEN.keys()),
            help="Selecciona el portafolio de credito"
        )

        # Campo para escribir manualmente - tiene prioridad sobre el selectbox
        cartera_manual = st.text_input(
            "O escribe el nombre de la cartera manualmente",
            value="",
            placeholder="Ej: Pyme, Tarjeta, Vehiculo, etc.",
            help="Si escribes aqui, se usara este valor en lugar del seleccionado arriba"
        )

        # Determinar cartera final: manual tiene prioridad
        if cartera_manual.strip():
            cartera_display = cartera_manual.strip()
            cartera = cartera_display.lower().replace(" ", "_").replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u").replace("ñ", "n")
        else:
            cartera = CARTERA_MAP_GEN[cartera_display]

    with col2:
        st.markdown(section_title("Modo Endogena"), unsafe_allow_html=True)
        modo_endo = st.radio(
            "Calculo de la endogena",
            options=["actual", "media_movil"],
            help="'actual': valor original | 'media_movil': promedio movil"
        )

        if modo_endo == "media_movil":
            ventana_mm = st.slider(
                "Ventana de media movil (meses)",
                min_value=1,
                max_value=12,
                value=3,
                help="Numero de meses para el calculo de promedio movil"
            )
        else:
            ventana_mm = 3

    st.markdown(divider(), unsafe_allow_html=True)
    st.markdown(section_title("Configuracion Generador"), unsafe_allow_html=True)

    col1, col2 = st.columns(2)

    with col1:
        usar_fechas_default = st.checkbox(
            "Usar fechas predeterminadas",
            value=True,
            help="Activar para usar: 2018-10-01 a 2025-03-01"
        )

        if usar_fechas_default:
            fecha_inicio = "2018-10-01"
            fecha_fin = "2025-03-01"
            st.info(f"Inicio: {fecha_inicio} | Fin: {fecha_fin}")
        else:
            col_a, col_b = st.columns(2)
            with col_a:
                fecha_inicio = st.text_input(
                    "Fecha inicio (YYYY-MM-DD)",
                    value="2018-10-01"
                )
            with col_b:
                fecha_fin = st.text_input(
                    "Fecha fin historico (YYYY-MM-DD)",
                    value="2025-03-01"
                )

    with col2:
        editar_nombres = st.checkbox(
            "Editar nombres de archivos",
            value=False,
            help="Si marcas esto, puedes personalizar los nombres"
        )

        if editar_nombres:
            st.warning("Cambiar el formato puede romper el flujo. Procede con cuidado.")

            # Generar nombres sugeridos basados en pais y cartera
            nombres_sugeridos = {
                "hist": f"hist_{pais.lower()}_{cartera.lower()}.xlsx",
                "base": f"base_{pais.lower()}_{cartera.lower()}.xlsx",
                "opt": f"opt_{pais.lower()}_{cartera.lower()}.xlsx",
                "adv": f"adv_{pais.lower()}_{cartera.lower()}.xlsx",
            }

            col_h, col_o = st.columns(2)
            with col_h:
                archivo_hist = st.text_input("Nombre archivo historico", value=nombres_sugeridos["hist"])
                archivo_base = st.text_input("Nombre archivo base", value=nombres_sugeridos["base"])
            with col_o:
                archivo_opt = st.text_input("Nombre archivo optimista", value=nombres_sugeridos["opt"])
                archivo_adv = st.text_input("Nombre archivo adverso", value=nombres_sugeridos["adv"])

            nombres_custom = {
                "hist": archivo_hist,
                "opt": archivo_opt,
                "base": archivo_base,
                "adv": archivo_adv,
            }
        else:
            nombres_custom = None

    st.markdown(divider(), unsafe_allow_html=True)
    st.markdown(section_title("Configuracion Motor SARIMAX"), unsafe_allow_html=True)

    col1, col2 = st.columns(2)

    with col1:
        tipo_modelo = st.selectbox(
            "Tipo de modelo",
            options=["total", "logit"],
            help="'total': escala original | 'logit': transformacion logit"
        )

    with col2:
        max_lags = st.slider(
            "Maximo de lags",
            min_value=1,
            max_value=12,
            value=8,
            help="Numero maximo de rezagos a evaluar en los modelos"
        )

    st.info("""
    Los siguientes parametros NO se editan aqui (requieren cambios en Colab):
    - Signos exogenas (positivo/negativo)
    - VIF maximo
    - Umbral de sensibilidad
    - Factor FWL (1.0 - 1.2)
    """)

    st.markdown(divider(), unsafe_allow_html=True)

    if st.button("Generar Notebooks", use_container_width=True, type="primary"):
        try:
            # Importar funciones del generador si existen
            try:
                from notebook_generator import (
                    generar_notebook_generador,
                    generar_notebook_motor,
                    generar_nombres_archivos,
                )
                gen_importado = True
            except ImportError:
                gen_importado = False
                st.error("No se encontro el modulo 'notebook_generator'. Verifica que este en el mismo directorio.")
                return

            # Generar nombres de archivos
            archivos = generar_nombres_archivos(pais, cartera)
            if nombres_custom:
                archivos = nombres_custom

            # Generar notebook del GENERADOR
            st.info("Generando notebook del Generador...")
            nb_gen = generar_notebook_generador(
                str(TEMPLATE_GENERADOR),
                pais=pais,
                cartera=cartera,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                modo_endo=modo_endo,
                ventana_mm=ventana_mm,
                nombres_custom=nombres_custom,
            )

            # Generar notebook del MOTOR
            st.info("Generando notebook del Motor...")
            nb_motor = generar_notebook_motor(
                str(TEMPLATE_MOTOR),
                pais=pais,
                cartera=cartera,
                tipo_modelo=tipo_modelo,
                max_lags=max_lags,
                nombres_archivos=archivos,
            )

            # Nombres de salida: usamos cartera_display para que refleje lo que el usuario selecciono
            cartera_slug = cartera.lower().replace(" ", "_")
            nb_gen_name = f"Generacion_Variacion_{pais}_{cartera_slug}.ipynb"
            nb_motor_name = f"Motor_Sarimax_{cartera_slug}_{pais}.ipynb"

            # Convertir notebooks a JSON (string) para descarga individual
            nb_gen_json = json.dumps(nb_gen, ensure_ascii=False, indent=1)
            nb_motor_json = json.dumps(nb_motor, ensure_ascii=False, indent=1)

            # Mostrar resumen
            st.success("Notebooks generados correctamente")

            st.markdown(section_title("Resumen de configuracion"), unsafe_allow_html=True)
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.markdown(card_kpi("Pais", pais), unsafe_allow_html=True)
            with col2:
                st.markdown(card_kpi("Cartera", cartera_display), unsafe_allow_html=True)
            with col3:
                st.markdown(card_kpi("Modelo", tipo_modelo), unsafe_allow_html=True)
            with col4:
                st.markdown(card_kpi("Max Lags", str(max_lags)), unsafe_allow_html=True)

            st.markdown(section_title("Descargar o abrir en Google Colab"), unsafe_allow_html=True)

            # --- NOTEBOOK GENERADOR ---
            st.markdown("<p style='font-size:14px;font-weight:600;color:{NAVY};margin:8px 0 4px;'>Generador de Variacion</p>".format(NAVY=NAVY), unsafe_allow_html=True)

            col_gen1, col_gen2 = st.columns(2)

            with col_gen1:
                st.download_button(
                    label=f"Descargar {nb_gen_name}",
                    data=nb_gen_json,
                    file_name=nb_gen_name,
                    mime="application/json",
                    use_container_width=True,
                )

            with col_gen2:
                nb_gen_b64 = base64.b64encode(nb_gen_json.encode("utf-8")).decode("utf-8")
                colab_url_gen = f"https://colab.research.google.com/notebook#data={nb_gen_b64}"
                st.link_button(
                    label="Abrir en Google Colab",
                    url=colab_url_gen,
                    use_container_width=True,
                )

            # --- NOTEBOOK MOTOR ---
            st.markdown("<p style='font-size:14px;font-weight:600;color:{NAVY};margin:8px 0 4px;'>Motor SARIMAX</p>".format(NAVY=NAVY), unsafe_allow_html=True)

            col_mot1, col_mot2 = st.columns(2)

            with col_mot1:
                st.download_button(
                    label=f"Descargar {nb_motor_name}",
                    data=nb_motor_json,
                    file_name=nb_motor_name,
                    mime="application/json",
                    use_container_width=True,
                )

            with col_mot2:
                nb_motor_b64 = base64.b64encode(nb_motor_json.encode("utf-8")).decode("utf-8")
                colab_url_motor = f"https://colab.research.google.com/notebook#data={nb_motor_b64}"
                st.link_button(
                    label="Abrir en Google Colab",
                    url=colab_url_motor,
                    use_container_width=True,
                )

        except Exception as e:
            st.error(f"Error al generar notebooks:\n{str(e)}")
            st.exception(e)

# =============================================================================
# SESSION STATE
# =============================================================================
_prefs_guardadas = cargar_prefs_sidebar()
for key, default in [
    ("uploaded_file", None), ("modelos_data", {}), ("meta_contexto", None),
    ("modelo_seleccionado", None),
    ("modelo_final", None),
    ("documento_metodologico_pdf", None),
    ("documento_metodologico_excel", None),
    ("documento_metodologico_pdf_nombre", None),
    ("documento_metodologico_excel_nombre", None),
    ("documento_metodologico_data", None),
    ("criterio_ordenamiento", _prefs_guardadas.get("criterio_ordenamiento", "Pruebas aprobadas ↓")),
    ("exog_sel", {}), ("pred_filtro", "Todas"),
    ("nav_sticky", _prefs_guardadas.get("nav_sticky", True)),
    ("pending_modelo", None),
    ("filtro_ljung", _prefs_guardadas.get("filtro_ljung", "Todos")),
    ("filtro_jarque", _prefs_guardadas.get("filtro_jarque", "Todos")),
    ("filtro_hetero", _prefs_guardadas.get("filtro_hetero", "Todos")),
    ("vista_resumen", True), ("comparar_sel", []),
    ("favoritos", set()),
    ("filtro_favoritos", _prefs_guardadas.get("filtro_favoritos", False)),
    ("vista_favoritos", False),
    ("sec_contexto", _prefs_guardadas.get("sec_contexto", True)),
    ("sec_orden", _prefs_guardadas.get("sec_orden", True)),
    ("sec_exogenas", _prefs_guardadas.get("sec_exogenas", True)),
    ("meta_contexto_manual", {}),
]:
    if key not in st.session_state:
        st.session_state[key] = default

# =============================================================================
# APP PRINCIPAL
# =============================================================================
st.set_page_config(page_title="SARIMAX IFRS 9", layout="wide")
inject_css()

# =========================================================================
# ATAJOS DE TECLADO (flechas para navegar, Escape para limpiar filtros)
# =========================================================================
with st.container(key="kb_hidden"):
    kb_c1, kb_c2, kb_c3 = st.columns(3)
    with kb_c1:
        kb_prev_clicked = st.button("KB_PREV", key="btn_kb_prev")
    with kb_c2:
        kb_next_clicked = st.button("KB_NEXT", key="btn_kb_next")
    with kb_c3:
        kb_reset_clicked = st.button("KB_RESET", key="btn_kb_reset")

if kb_reset_clicked:
    st.session_state.filtro_ljung_sel = "Todos"
    st.session_state.filtro_jarque_sel = "Todos"
    st.session_state.filtro_hetero_sel = "Todos"
    st.session_state.filtro_ljung = "Todos"
    st.session_state.filtro_jarque = "Todos"
    st.session_state.filtro_hetero = "Todos"
    st.rerun()

if (kb_prev_clicked or kb_next_clicked) and st.session_state.modelos_data and st.session_state.modelo_seleccionado:
    modelos_list_kb, _, _, _ = construir_opciones_modelos()
    if st.session_state.modelo_seleccionado in modelos_list_kb:
        idx_kb = modelos_list_kb.index(st.session_state.modelo_seleccionado)
        if kb_prev_clicked and idx_kb > 0:
            st.session_state.pending_modelo = modelos_list_kb[idx_kb - 1]
            st.rerun()
        elif kb_next_clicked and idx_kb < len(modelos_list_kb) - 1:
            st.session_state.pending_modelo = modelos_list_kb[idx_kb + 1]
            st.rerun()

_KB_SHORTCUT_HTML = """
<script>
(function() {
    if (window.parent.__kbListenerAdded) { return; }
    window.parent.__kbListenerAdded = true;
    function clickHiddenButton(cssKey, fallbackText) {
        const doc = window.parent.document;
        let btn = doc.querySelector('.st-key-' + cssKey + ' button');
        if (!btn) {
            const all = doc.querySelectorAll('button');
            for (const b of all) {
                if (b.innerText && b.innerText.trim() === fallbackText) { btn = b; break; }
            }
        }
        if (btn) { btn.click(); }
    }
    window.parent.document.addEventListener('keydown', function(e) {
        const tag = (e.target && e.target.tagName) || '';
        if (['INPUT', 'TEXTAREA', 'SELECT'].includes(tag)) return;
        if (e.key === 'ArrowLeft') {
            clickHiddenButton('btn_kb_prev', 'KB_PREV');
        } else if (e.key === 'ArrowRight') {
            clickHiddenButton('btn_kb_next', 'KB_NEXT');
        } else if (e.key === 'Escape') {
            clickHiddenButton('btn_kb_reset', 'KB_RESET');
        }
    });
})();
</script>
"""
if hasattr(st, "iframe"):
    st.iframe(_KB_SHORTCUT_HTML, height=1, width=1)
else:
    components.html(_KB_SHORTCUT_HTML, height=0, width=0)

# =========================================================================
# HEADER Y PESTANAS
# =========================================================================
st.markdown(f"""
<div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:4px;">
    <div>
        <p style="font-size:22px;font-weight:700;color:{NAVY};margin:0;">SARIMAX IFRS 9</p>
        <p style="font-size:12px;color:{MUTED};margin:4px 0 0;">Generador y Dashboard de Modelos</p>
    </div>
</div>
<div style="height:1px;background:{BORDER};margin:12px 0 16px;"></div>
""", unsafe_allow_html=True)

tab_gen, tab_dash, tab_concat = st.tabs(["Generador", "Dashboard", "Concatenador"])

# =========================================================================
# TAB: GENERADOR
# =========================================================================
with tab_gen:
    render_generador()

# =========================================================================
# TAB: DASHBOARD
# =========================================================================
with tab_dash:
    col_left, col_right = st.columns([1, 4])

    # --- SIDEBAR DEL DASHBOARD ---
    with col_left:
        st.markdown(f"<p style='font-size:12px;font-weight:700;color:{NAVY};margin:0 0 10px;letter-spacing:0.5px;'>CARGAR MODELO</p>", unsafe_allow_html=True)
        uploaded = st.file_uploader("Archivo Excel (.xlsx)", type=["xlsx"], label_visibility="collapsed")
        if uploaded is not None:
            st.session_state.uploaded_file = uploaded
            if not st.session_state.modelos_data or uploaded.name != getattr(st.session_state, 'last_file_name', None):
                with st.spinner("Parseando modelos..."):
                    st.session_state.modelos_data = parsear_excel(uploaded)
                    st.session_state.meta_contexto = leer_meta_embebida(uploaded)
                    # Si el nuevo archivo tiene metadata embebida, limpiar fallback manual
                    if st.session_state.meta_contexto is not None:
                        st.session_state.meta_contexto_manual = {}
                    st.session_state.last_file_name = uploaded.name
                    st.session_state.vista_resumen = True
                    st.session_state.comparar_sel = []
                    st.session_state.modelo_final = None
                    st.session_state.documento_metodologico_pdf = None
                    st.session_state.documento_metodologico_excel = None
                    st.session_state.documento_metodologico_pdf_nombre = None
                    st.session_state.documento_metodologico_excel_nombre = None
                    st.session_state.documento_metodologico_data = None
                st.success(f"Archivo cargado: {uploaded.name}")
                if st.session_state.modelo_seleccionado is None or st.session_state.modelo_seleccionado not in st.session_state.modelos_data:
                    st.session_state.modelo_seleccionado = list(st.session_state.modelos_data.keys())[0]
        if st.session_state.uploaded_file is not None:
            if st.button("Eliminar archivo", key="btn_eliminar", use_container_width=True):
                st.session_state.uploaded_file = None
                st.session_state.modelos_data = {}
                st.session_state.meta_contexto = None
                st.session_state.modelo_seleccionado = None
                st.session_state.last_file_name = None
                st.session_state.exog_sel = {}
                st.session_state.vista_resumen = True
                st.session_state.comparar_sel = []
                st.session_state.modelo_final = None
                st.session_state.documento_metodologico_pdf = None
                st.session_state.documento_metodologico_excel = None
                st.session_state.documento_metodologico_pdf_nombre = None
                st.session_state.documento_metodologico_excel_nombre = None
                st.session_state.documento_metodologico_data = None
                st.rerun()
        if st.session_state.modelos_data:
            st.markdown(divider(), unsafe_allow_html=True)
            meta = st.session_state.meta_contexto
            meta_fallback = st.session_state.get("meta_contexto_manual", {})
            meta_display = meta if meta else meta_fallback

            # Formulario manual de metadata (solo si no hay metadata embebida)
            if not meta:
                with st.expander("Completar metadata manualmente", expanded=not bool(meta_fallback)):
                    _pais = st.selectbox("País", ["Colombia", "Panama", "Costa Rica"], key="manual_pais")
                    _cartera = st.text_input("Cartera", key="manual_cartera")
                    _tipo_endogena = st.selectbox("Tipo de endógena", ["total", "logit"], key="manual_tipo")
                    _modo_endogena = st.selectbox("Modo de endógena", ["actual", "media_movil"], key="manual_modo")
                    _ventana_mm = st.number_input("Ventana media móvil (meses)", 1, 12, 3, key="manual_ventana")
                    _vif_max = st.number_input("VIF máximo", 1.0, 20.0, 10.0, key="manual_vif")
                    _fwl_min = st.number_input("FWL mínimo", 0.0, 2.0, 1.0, key="manual_fwl_min")
                    _fwl_max = st.number_input("FWL máximo", 0.0, 2.0, 1.2, key="manual_fwl_max")
                    _valores_p = st.text_input("Órdenes AR candidatos", value="[0,1,2,3]", key="manual_p")
                    _valores_q = st.text_input("Órdenes MA candidatos", value="[0,1,2,3]", key="manual_q")
                    _max_lags = st.number_input("Máximo lags", 1, 12, 8, key="manual_lags")
                    _umbral_sens = st.number_input("Umbral sensibilidad", 0.0, 1.0, 0.00005, format="%.5f", key="manual_sens")
                    _trend = st.text_input("Trend del modelo", value="c", key="manual_trend")

                    if st.button("Aplicar metadata", use_container_width=True, key="btn_aplicar_manual"):
                        st.session_state.meta_contexto_manual = {
                            "pais": _pais,
                            "cartera": _cartera,
                            "motor_tipo_endogena": _tipo_endogena,
                            "generador_modo_endogena": _modo_endogena,
                            "generador_ventana_mm": _ventana_mm,
                            "motor_vif_max": _vif_max,
                            "motor_fwl_factor_min": _fwl_min,
                            "motor_fwl_factor_max": _fwl_max,
                            "motor_valores_p": _valores_p,
                            "motor_valores_q": _valores_q,
                            "motor_max_lags": _max_lags,
                            "motor_umbral_sensibilidad": _umbral_sens,
                            "motor_trend": _trend,
                            "motor_top_exportar": 5,
                            "motor_max_exog_por_modelo": 4,
                        }
                        st.success("Metadata manual aplicada.")
                        st.rerun()

            if encabezado_colapsable("Contexto de la corrida", "sec_contexto"):
                if meta_display:
                    meta_kpis = extraer_kpis_meta(meta, meta_fallback)
                    c1, c2 = st.columns(2)
                    with c1:
                        pais_nombre = meta_kpis.get('pais', '-')
                        codigo_iso = BANDERAS_PAISES.get(pais_nombre, pais_nombre).upper()
                        bandera_html = obtener_bandera_pais(pais_nombre)
                        valor_pais = f"{bandera_html}{codigo_iso}"
                        st.markdown(card_kpi("Pais", valor_pais), unsafe_allow_html=True)
                    with c2:
                        st.markdown(card_kpi("Ventana media movil", meta_kpis.get('ventana_mm', '-')), unsafe_allow_html=True)
                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown(card_kpi("Cartera", meta_kpis.get('cartera', '-'), accent=BLUE), unsafe_allow_html=True)
                    with c2:
                        fwl_range = f"{meta_kpis.get('fwl_min', '?')} - {meta_kpis.get('fwl_max', '?')}"
                        st.markdown(card_kpi("Rango FWL", fwl_range, accent=GREEN), unsafe_allow_html=True)
                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown(card_kpi("Modo endogena", meta_kpis.get('modo_endogena', '-')), unsafe_allow_html=True)
                    with c2:
                        tipo_endog = meta_kpis.get('tipo_endogena', '-')
                        # Mostrar en mayusculas y con color si es logit
                        if tipo_endog == 'logit':
                            tipo_display = f"<span style='color:#1E5AA8;font-weight:700;'>{tipo_endog.upper()}</span>"
                        elif tipo_endog == 'total':
                            tipo_display = f"<span style='color:#1A6B3E;font-weight:700;'>{tipo_endog.upper()}</span>"
                        else:
                            tipo_display = tipo_endog
                        st.markdown(card_kpi("Tipo endogena", tipo_display), unsafe_allow_html=True)
                else:
                    st.caption("Sin metadata embebida ni manual.")
            st.markdown(divider(), unsafe_allow_html=True)
            if encabezado_colapsable("Ordenar y Filtrar", "sec_orden"):
                criterio = st.radio("Ordenar por:", ["Nombre (A-Z)", "Pruebas aprobadas ↓", "Pruebas aprobadas ↑",
                                                      "Score global ↓", "Score global ↑"],
                                    index=["Nombre (A-Z)", "Pruebas aprobadas ↓", "Pruebas aprobadas ↑",
                                           "Score global ↓", "Score global ↑"].index(st.session_state.criterio_ordenamiento),
                                    key="criterio_orden")
                st.session_state.criterio_ordenamiento = criterio
                st.markdown(f"<p style='font-size:10px;color:{MUTED};margin:8px 0 4px;'>Ljung-Box</p>", unsafe_allow_html=True)
                filtro_ljung = st.selectbox("Ljung-Box", ["Todos", "A o B (Cumple)", "A, B o C", "Solo A"],
                                             index=["Todos", "A o B (Cumple)", "A, B o C", "Solo A"].index(st.session_state.filtro_ljung),
                                             key="filtro_ljung_sel", label_visibility="collapsed")
                st.session_state.filtro_ljung = filtro_ljung
                st.markdown(f"<p style='font-size:10px;color:{MUTED};margin:8px 0 4px;'>Jarque-Bera</p>", unsafe_allow_html=True)
                filtro_jarque = st.selectbox("Jarque-Bera", ["Todos", "A o B (Cumple)", "A, B o C", "Solo A"],
                                              index=["Todos", "A o B (Cumple)", "A, B o C", "Solo A"].index(st.session_state.filtro_jarque),
                                              key="filtro_jarque_sel", label_visibility="collapsed")
                st.session_state.filtro_jarque = filtro_jarque
                st.markdown(f"<p style='font-size:10px;color:{MUTED};margin:8px 0 4px;'>Heterocedasticidad</p>", unsafe_allow_html=True)
                filtro_hetero = st.selectbox("Heterocedasticidad", ["Todos", "A o B (Cumple)", "A, B o C", "Solo A"],
                                              index=["Todos", "A o B (Cumple)", "A, B o C", "Solo A"].index(st.session_state.filtro_hetero),
                                              key="filtro_hetero_sel", label_visibility="collapsed")
                st.session_state.filtro_hetero = filtro_hetero
                st.markdown("<div style='height:6px;'></div>", unsafe_allow_html=True)
                filtro_favoritos = st.checkbox("Solo favoritos", key="filtro_favoritos_sel",
                                                value=st.session_state.get("filtro_favoritos", False))
                st.session_state.filtro_favoritos = filtro_favoritos
            st.markdown(divider(), unsafe_allow_html=True)
            modelos_list, pruebas_dict, scores_dict, global_dict = construir_opciones_modelos()
            if st.session_state.pending_modelo is not None and st.session_state.pending_modelo in modelos_list:
                st.session_state.modelo_seleccionado = st.session_state.pending_modelo
                st.session_state["sel_modelo"] = label_modelo(st.session_state.pending_modelo, pruebas_dict, scores_dict, global_dict)
                st.session_state.pending_modelo = None
            opciones = [label_modelo(m, pruebas_dict, scores_dict, global_dict) for m in modelos_list]
            if not modelos_list:
                st.warning("Ningun modelo cumple con los filtros seleccionados.")
                st.session_state.modelo_seleccionado = None
            else:
                idx = modelos_list.index(st.session_state.modelo_seleccionado) if st.session_state.modelo_seleccionado in modelos_list else 0
                seleccion = st.selectbox("Modelo", opciones, index=idx, key="sel_modelo")
                nombre_parseado = seleccion.split("  (")[0]
                if nombre_parseado.startswith("[F] "):
                    nombre_parseado = nombre_parseado[4:]
                st.session_state.modelo_seleccionado = nombre_parseado
            st.markdown(divider(), unsafe_allow_html=True)
            st.toggle("Fijar flechas de navegacion", key="nav_sticky",
                      help="Mantiene los botones Anterior/Siguiente siempre visibles, flotando sobre la pagina al hacer scroll.")
            guardar_prefs_sidebar()
            if st.session_state.modelo_seleccionado:
                datos = st.session_state.modelos_data.get(st.session_state.modelo_seleccionado, {})
                st.markdown(f"<p style='font-size:11px;font-weight:600;color:{NAVY};margin:12px 0 4px;'>MODELO ACTUAL</p>", unsafe_allow_html=True)
                st.markdown(f"<p style='font-size:14px;font-weight:700;color:{NAVY};margin:0;'>{st.session_state.modelo_seleccionado}</p>", unsafe_allow_html=True)
                st.markdown(f"<p style='font-size:11px;color:{MUTED};margin:4px 0 0;'>{datos.get('observaciones', 0)} observaciones</p>", unsafe_allow_html=True)
                st.markdown("<div style='height:6px;'></div>", unsafe_allow_html=True)
                boton_favorito(st.session_state.modelo_seleccionado, key_suffix="sidebar")
                exogenas = datos.get('exogenas_nombres', [])
                if exogenas:
                    st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)
                    if encabezado_colapsable("Exogenas", "sec_exogenas"):
                        coefs = datos.get('coeficientes')
                        sigs = obtener_significancia_exogenas(coefs, exogenas)
                        sig_count = sum(1 for _, _, s in sigs if s == "Significativa")
                        st.markdown(f"<p style='font-size:10px;color:{MUTED};margin:0 0 6px;'>{sig_count} de {len(exogenas)} significativas</p>", unsafe_allow_html=True)
                        for ex, pval, status in sigs:
                            color = GREEN if status == "Significativa" else (RED if status == "No significativa" else "#B8860B")
                            label = "SIG" if status == "Significativa" else ("NO SIG" if status == "No significativa" else "MARG")
                            p_txt = f"p={pval:.3f}" if pval is not None else "p=N/A"
                            st.markdown(f'<div style="display:flex;justify-content:space-between;align-items:center;padding:2px 0;font-size:11px;"><span style="color:{TEXT}">{ex}</span><span style="color:{color};font-weight:600;">{label} ({p_txt})</span></div>', unsafe_allow_html=True)

    # --- PANEL PRINCIPAL DEL DASHBOARD ---
    with col_right:
        if not st.session_state.modelos_data:
            st.markdown(f"""
            <div style="background:{WHITE};border:1px solid {BORDER};border-radius:10px;padding:60px 32px;text-align:center;margin-top:40px;">
                <p style="font-size:18px;font-weight:700;color:{NAVY};margin:0 0 8px;">Dashboard SARIMAX</p>
                <p style="font-size:13px;color:{MUTED};margin:0;">Suba un archivo Excel para comenzar el analisis.</p>
            </div>
            """, unsafe_allow_html=True)
        elif st.session_state.vista_favoritos:
            render_vista_favoritos()
        elif st.session_state.vista_resumen:
            render_resumen_ejecutivo()
        elif st.session_state.modelo_seleccionado is None:
            st.markdown(f"""
            <div style="background:{WHITE};border:1px solid {BORDER};border-radius:10px;padding:60px 32px;text-align:center;margin-top:40px;">
                <p style="font-size:18px;font-weight:700;color:{NAVY};margin:0 0 8px;">Sin modelos disponibles</p>
                <p style="font-size:13px;color:{MUTED};margin:0;">Ajuste los filtros de diagnostico para ver modelos.</p>
            </div>
            """, unsafe_allow_html=True)
        else:
            datos = st.session_state.modelos_data.get(st.session_state.modelo_seleccionado, {})
            meta_kpis = extraer_kpis_meta(st.session_state.meta_contexto, st.session_state.get("meta_contexto_manual", {}))
            pais = meta_kpis.get('pais', '-')
            cartera = meta_kpis.get('cartera', '-')
            pais_codigo_hdr = BANDERAS_PAISES.get(pais, pais).upper()
            score_global_hdr, _ = calcular_score_global(datos.get('pruebas'))
            st.markdown(f"""
            <div style="display:flex;align-items:flex-end;gap:16px;margin-bottom:4px;">
                <div style="flex:1;">
                    <p style="font-size:11px;color:{MUTED};margin:0;text-transform:uppercase;letter-spacing:0.5px;font-weight:600;">Modelo seleccionado</p>
                    <p style="font-size:20px;font-weight:700;color:{NAVY};margin:4px 0 0;">{st.session_state.modelo_seleccionado}
                        <span style="margin-left:10px;">{score_global_badge(score_global_hdr)}</span>
                    </p>
                </div>
                <div style="text-align:right;">
                    <p style="font-size:11px;color:{MUTED};margin:0;">{obtener_bandera_pais(pais)}{pais_codigo_hdr} - {cartera} | {meta_kpis.get('tipo_endogena', '').upper()}</p>
                    <p style="font-size:11px;color:{LTGRAY};margin:2px 0 0;">{len(st.session_state.modelos_data)} modelos cargados</p>
                </div>
            </div>
            <div style="height:1px;background:{BORDER};margin:12px 0 16px;"></div>
            """, unsafe_allow_html=True)
            if st.session_state.get("modelo_final") == st.session_state.modelo_seleccionado:
                st.markdown(
                    f'<div style="display:inline-block;background:#fff4d6;color:{NAVY};'
                    f'border:1px solid #f0d58a;border-radius:6px;padding:6px 10px;'
                    f'font-size:12px;font-weight:700;margin:0 0 10px;"> MODELO FINAL SELECCIONADO</div>',
                    unsafe_allow_html=True
                )
            elif st.session_state.get("modelo_final"):
                st.markdown(
                    f"<p style='font-size:11px;color:{MUTED};margin:0 0 10px;'>"
                    f"Modelo final actual: <b>{st.session_state.get('modelo_final')}</b></p>",
                    unsafe_allow_html=True
                )

            hcol1, hcol2, hcol3, hcol4 = st.columns([1.2, 1, 1, 1.3])
            with hcol1:
                if st.button("Ver resumen de la corrida", key="btn_ver_resumen", use_container_width=True):
                    st.session_state.vista_resumen = True
                    st.rerun()
            with hcol2:
                n_fav = len([m for m in st.session_state.get("favoritos", set()) if m in st.session_state.modelos_data])
                if st.button(f"Ver favoritos ({n_fav})", key="btn_ver_favoritos_detalle", use_container_width=True):
                    st.session_state.vista_favoritos = True
                    st.rerun()
            with hcol3:
                boton_favorito(st.session_state.modelo_seleccionado, key_suffix="detalle")
            with hcol4:
                if st.button(" Marcar como modelo final", key="btn_modelo_final_top", use_container_width=True):
                    with st.spinner("Generando documento metodológico..."):
                        ok = generar_y_guardar_documentos(st.session_state.modelo_seleccionado)
                    if ok:
                        st.success("✅ Documento generado. Descarga abajo.")
                    st.rerun()
            # =====================================================================
            modelos_list, pruebas_dict_nav, scores_dict_nav, global_dict_nav = construir_opciones_modelos()
            current_idx = modelos_list.index(st.session_state.modelo_seleccionado) if st.session_state.modelo_seleccionado in modelos_list else 0
            tab_resumen, tab1, tab2, tab3, tab4 = st.tabs(["Resumen Modelo", "Visualizacion", "Predicciones", "Diagnosticos", "Comparar"])
            # TAB 0: RESUMEN MODELO
            # =====================================================================
            with tab_resumen:
                st.markdown(section_title("Factor FWL a 12 meses"), unsafe_allow_html=True)
                df_fwl_resumen = datos.get('fwl_12m')
                if df_fwl_resumen is not None and not df_fwl_resumen.empty:
                    st.plotly_chart(fig_fwl_12m(df_fwl_resumen), use_container_width=True, key="resumen_fwl12m")
                else:
                    st.info("No hay datos de FWL a 12 meses.")
                st.markdown(divider(), unsafe_allow_html=True)
                render_diagnosticos_corporativo(datos.get('pruebas'), mostrar_detalle_tecnico=False)
                st.markdown(divider(), unsafe_allow_html=True)
                render_seccion_coeficientes(datos, key_prefix="resumen")
            # =====================================================================
            # TAB 1: VISUALIZACION
            # =====================================================================
            with tab1:
                st.markdown(section_title("Exogenas activas"), unsafe_allow_html=True)
                exogenas = datos.get('exogenas_nombres', [])
                modelo_key = st.session_state.modelo_seleccionado
                if modelo_key not in st.session_state.exog_sel:
                    st.session_state.exog_sel[modelo_key] = []
                if exogenas:
                    n_cols = min(len(exogenas), 6)
                    chip_cols = st.columns(n_cols)
                    for i, ex in enumerate(exogenas):
                        with chip_cols[i % n_cols]:
                            activo = ex in st.session_state.exog_sel[modelo_key]
                            label = f"[x] {ex}" if activo else f"[ ] {ex}"
                            if st.button(label, key=f"chip_{modelo_key}_{ex}", use_container_width=True):
                                if activo:
                                    st.session_state.exog_sel[modelo_key].remove(ex)
                                else:
                                    st.session_state.exog_sel[modelo_key].append(ex)
                                st.rerun()
                else:
                    st.caption("Sin exogenas en este modelo.")
                df_end = datos.get('fecha_endogena')
                endogena_cols = datos.get('endogenas_cols', [])
                if df_end is not None and not df_end.empty and endogena_cols:
                    fig = fig_predicciones(df_end, endogena_cols, datos.get('exogenas'), st.session_state.exog_sel.get(modelo_key, []), st.session_state.modelo_seleccionado)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("No hay datos de predicciones.")
                st.markdown(divider(), unsafe_allow_html=True)
                st.markdown(section_title("Factor FWL por ano y escenario"), unsafe_allow_html=True)
                df_fwl_anual = datos.get('fwl_anual')
                if df_fwl_anual is not None and not df_fwl_anual.empty:
                    try:
                        df_pivot = df_fwl_anual.pivot(index='Ano', columns='Escenario', values='Factor FWL').reset_index()
                        rename_map = {}
                        for c in df_pivot.columns:
                            c_str = str(c).lower()
                            if 'base' in c_str: rename_map[c] = 'Base'
                            elif 'adverso' in c_str or 'advers' in c_str: rename_map[c] = 'Adverso'
                            elif 'optimista' in c_str: rename_map[c] = 'Optimista'
                        df_pivot = df_pivot.rename(columns=rename_map)
                        st.dataframe(df_pivot, use_container_width=True, hide_index=True)
                    except:
                        st.dataframe(df_fwl_anual, use_container_width=True, hide_index=True)
                else:
                    st.info("No hay datos de Factor FWL por Ano.")
                st.markdown(divider(), unsafe_allow_html=True)
                st.markdown(section_title("Factor FWL a 12 meses"), unsafe_allow_html=True)
                df_fwl = datos.get('fwl_12m')
                if df_fwl is not None and not df_fwl.empty:
                    st.plotly_chart(fig_fwl_12m(df_fwl), use_container_width=True, key="viz_fwl12m")
                else:
                    st.info("No hay datos de FWL a 12 meses.")
                st.markdown(divider(), unsafe_allow_html=True)
                st.markdown(section_title("Factor FWL ponderado"), unsafe_allow_html=True)
                c1, c2, c3 = st.columns(3)
                with c1: peso_base = st.number_input("Peso Base", 0.0, 1.0, 0.33, 0.01, key="pw_base")
                with c2: peso_adverso = st.number_input("Peso Adverso", 0.0, 1.0, 0.33, 0.01, key="pw_adv")
                with c3: peso_optimista = st.number_input("Peso Optimista", 0.0, 1.0, 0.34, 0.01, key="pw_opt")
                suma = peso_base + peso_adverso + peso_optimista
                if abs(suma - 1.0) < 0.001:
                    st.markdown(pill("VALIDO", "#E8F5E9", GREEN), unsafe_allow_html=True)
                elif suma < 1.0:
                    st.markdown(pill(f"{1.0-suma:.2f} DISPONIBLE", "#FFF8E1", "#B8860B"), unsafe_allow_html=True)
                else:
                    st.markdown(pill(f"EXCEDE {suma-1.0:.2f}", "#FFEBEE", RED), unsafe_allow_html=True)
                if df_fwl is not None and not df_fwl.empty and suma <= 1.0:
                    pesos = {'base': peso_base, 'adverso': peso_adverso, 'optimista': peso_optimista}
                    df_pond = calcular_fwl_ponderado(df_fwl, pesos)
                    if df_pond is not None:
                        res = resumen_fwl(df_pond)
                        if res:
                            c1, c2, c3, c4 = st.columns(4)
                            with c1: st.markdown(card_metric("Promedio", f"{res.get('promedio', 0):.4f}", BLUE), unsafe_allow_html=True)
                            with c2: st.markdown(card_metric("Maximo", f"{res.get('maximo', 0):.4f}", GREEN), unsafe_allow_html=True)
                            with c3: st.markdown(card_metric("Minimo", f"{res.get('minimo', 0):.4f}", RED), unsafe_allow_html=True)
                            with c4: st.markdown(card_metric("Volatilidad (s)", f"{res.get('volatilidad', 0):.4f}", GRAY), unsafe_allow_html=True)
                        st.plotly_chart(fig_fwl_ponderado(df_pond), use_container_width=True)
                    else:
                        st.info("No se pudo calcular el FWL ponderado.")
                elif suma > 1.0:
                    st.warning("Ajuste los pesos para que la suma no exceda 1.0")
            # =====================================================================
            # TAB 2: PREDICCIONES
            # =====================================================================
            with tab2:
                st.markdown(section_title("Datos de prediccion"), unsafe_allow_html=True)
                filtros = st.columns(4)
                with filtros[0]:
                    if st.button("Ver Base", use_container_width=True): st.session_state.pred_filtro = "Base"
                with filtros[1]:
                    if st.button("Ver Adverso", use_container_width=True): st.session_state.pred_filtro = "Adverso"
                with filtros[2]:
                    if st.button("Ver Optimista", use_container_width=True): st.session_state.pred_filtro = "Optimista"
                with filtros[3]:
                    if st.button("Ver todas", use_container_width=True): st.session_state.pred_filtro = "Todas"
                st.markdown(f"<p style='font-size:11px;color:{MUTED};margin:8px 0;'>Filtro activo: <b>{st.session_state.pred_filtro}</b></p>", unsafe_allow_html=True)
                df_end = datos.get('fecha_endogena')
                endogena_cols = datos.get('endogenas_cols', [])
                if df_end is not None and not df_end.empty and endogena_cols:
                    fecha_col = 'fecha' if 'fecha' in df_end.columns else df_end.columns[0]
                    base_col = adv_col = opt_col = None
                    for col in endogena_cols:
                        col_str = str(col).upper()
                        if col_str == 'BASE': base_col = col
                        elif col_str in ['ADVERSO', 'ADVERSA']: adv_col = col
                        elif col_str == 'OPTIMISTA': opt_col = col
                    df_pred = pd.DataFrame()
                    df_pred['Fecha'] = pd.to_datetime(df_end[fecha_col]).dt.strftime('%Y-%m-%d')
                    cols_export = ['Fecha']
                    if base_col and base_col in df_end.columns and st.session_state.pred_filtro in ["Base", "Todas"]:
                        df_pred['Base'] = df_end[base_col].astype(float).round(4)
                        cols_export.append('Base')
                    if adv_col and adv_col in df_end.columns and st.session_state.pred_filtro in ["Adverso", "Todas"]:
                        df_pred['Adverso'] = df_end[adv_col].astype(float).round(4)
                        cols_export.append('Adverso')
                    if opt_col and opt_col in df_end.columns and st.session_state.pred_filtro in ["Optimista", "Todas"]:
                        df_pred['Optimista'] = df_end[opt_col].astype(float).round(4)
                        cols_export.append('Optimista')
                    st.dataframe(df_pred[cols_export], use_container_width=True, hide_index=True, height=400)
                    csv = df_pred[cols_export].to_csv(index=False).encode('utf-8')
                    st.download_button("Descargar CSV", csv, f"predicciones_{st.session_state.modelo_seleccionado}.csv", "text/csv")
                else:
                    st.info("No hay datos de predicciones.")
            # =====================================================================
            # TAB 3: DIAGNOSTICOS
            # =====================================================================
            with tab3:
                st.markdown(render_leyenda_scores(), unsafe_allow_html=True)
                pruebas = datos.get('pruebas')
                render_metricas_diagnostico(pruebas)
                st.markdown(divider(), unsafe_allow_html=True)
                render_diagnosticos_corporativo(pruebas)
                st.markdown(divider(), unsafe_allow_html=True)
                st.markdown(section_title("Distribucion de residuos"), unsafe_allow_html=True)
                residuos = datos.get('residuos_ind')
                if residuos is not None and not residuos.empty:
                    res_col = None
                    for c in residuos.columns:
                        if 'residuo' in str(c).lower():
                            res_col = c
                            break
                    if res_col:
                        vals = residuos[res_col].dropna().astype(float)
                        media, std = vals.mean(), vals.std()
                        st.plotly_chart(fig_histograma_residuos(vals, media, std), use_container_width=True)
                        st.markdown(section_title("Estadisticas descriptivas"), unsafe_allow_html=True)
                        c1, c2, c3, c4, c5 = st.columns(5)
                        with c1: st.markdown(card_metric("Media", f"{media:.4f}"), unsafe_allow_html=True)
                        with c2: st.markdown(card_metric("Desv. Std.", f"{std:.4f}"), unsafe_allow_html=True)
                        with c3: st.markdown(card_metric("Asimetria", f"{stats.skew(vals):.4f}"), unsafe_allow_html=True)
                        with c4: st.markdown(card_metric("Curtosis", f"{stats.kurtosis(vals):.4f}"), unsafe_allow_html=True)
                        with c5: st.markdown(card_metric("Observaciones", f"{len(vals)}"), unsafe_allow_html=True)
                else:
                    st.info("No hay datos de residuos.")
                st.markdown(divider(), unsafe_allow_html=True)
                render_seccion_coeficientes(datos, key_prefix="diag")
            # =====================================================================
            # TAB 4: COMPARAR
            # =====================================================================
            with tab4:
                st.markdown(section_title("Comparador de modelos"), unsafe_allow_html=True)
                st.markdown(f"<p style='font-size:11px;color:{MUTED};margin:0 0 10px;'>Seleccione hasta 3 modelos para comparar lado a lado.</p>", unsafe_allow_html=True)
                todos_modelos = list(st.session_state.modelos_data.keys())
                default_sel = [m for m in st.session_state.get("comparar_sel", []) if m in todos_modelos][:3]
                seleccion_comp = st.multiselect("Modelos a comparar", todos_modelos, default=default_sel, key="comparar_multiselect")
                if len(seleccion_comp) > 3:
                    st.warning("Se seleccionaron mas de 3 modelos. Solo se compararan los primeros 3.")
                    seleccion_comp = seleccion_comp[:3]
                st.session_state.comparar_sel = seleccion_comp
                if seleccion_comp:
                    cols_comp = st.columns(len(seleccion_comp))
                    for i_comp, nombre_comp in enumerate(seleccion_comp):
                        with cols_comp[i_comp]:
                            render_columna_comparacion(nombre_comp)
                else:
                    st.info("Seleccione al menos un modelo para iniciar la comparacion.")
            st.markdown(divider(), unsafe_allow_html=True)
            cfinal1, cfinal2, cfinal3 = st.columns([1.3, 1, 1])
            with cfinal1:
                if st.button(" Marcar como modelo final", key="btn_modelo_final_bottom", use_container_width=True):
                    with st.spinner("Generando documento metodológico..."):
                        ok = generar_y_guardar_documentos(st.session_state.modelo_seleccionado)
                    if ok:
                        st.success("✅ Documento generado. Descarga abajo.")
                    st.rerun()
            with cfinal2:
                st.download_button(
                    "⬇️ Descargar PDF",
                    data=st.session_state.documento_metodologico_pdf or b"",
                    file_name=st.session_state.documento_metodologico_pdf_nombre or "Documento_Metodologico.pdf",
                    mime="application/pdf",
                    disabled=not bool(st.session_state.documento_metodologico_pdf),
                    use_container_width=True
                )
            with cfinal3:
                st.download_button(
                    "⬇️ Descargar Excel",
                    data=st.session_state.documento_metodologico_excel or b"",
                    file_name=st.session_state.documento_metodologico_excel_nombre or "Documento_Metodologico.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    disabled=not bool(st.session_state.documento_metodologico_excel),
                    use_container_width=True
                )

            # --- Bottom nav bar ---
            nav_sticky = st.session_state.get("nav_sticky", True)
            if nav_sticky:
                st.markdown(f"""
                <style>
                div[data-testid="stVerticalBlockBorderWrapper"]:has(> div > div.st-key-nav_flechas),
                .st-key-nav_flechas {{
                    position: fixed !important;
                    bottom: 22px;
                    left: 50%;
                    transform: translateX(-46%);
                    z-index: 9999;
                    background: {WHITE};
                    border: 1px solid {BORDER};
                    border-radius: 14px;
                    box-shadow: 0 8px 28px rgba(11,37,69,0.16);
                    padding: 6px 10px !important;
                    width: auto !important;
                    max-width: 560px;
                }}
                .block-container {{ padding-bottom: 120px !important; }}
                </style>
                """, unsafe_allow_html=True)
            else:
                st.markdown("<div style='height:40px;'></div>", unsafe_allow_html=True)
            with st.container(key="nav_flechas"):
                nav_cols = st.columns([1, 2, 1])
                with nav_cols[0]:
                    if st.button("Anterior", disabled=current_idx == 0, key="btn_prev_real", use_container_width=True):
                        st.session_state.pending_modelo = modelos_list[current_idx - 1]
                        st.rerun()
                with nav_cols[1]:
                    st.markdown(
                        f"""
                        <div style="text-align:center;padding:6px 4px;">
                            <p style="font-size:10px;color:{MUTED};margin:0;text-transform:uppercase;letter-spacing:0.5px;font-weight:600;">Modelo {current_idx + 1} de {len(modelos_list)}</p>
                            <p style="font-size:13px;color:{NAVY};font-weight:700;margin:2px 0 0;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">{st.session_state.modelo_seleccionado}</p>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )
                with nav_cols[2]:
                    if st.button("Siguiente", disabled=current_idx == len(modelos_list) - 1, key="btn_next_real", use_container_width=True):
                        st.session_state.pending_modelo = modelos_list[current_idx + 1]
                        st.rerun()

# =========================================================================
# TAB: CONCATENADOR
# =========================================================================
with tab_concat:
    from ui_concatenador import render_concatenador

    render_concatenador(
        NAVY=NAVY,
        BLUE=BLUE,
        GREEN=GREEN,
        RED=RED,
        GRAY=GRAY,
        LTGRAY=LTGRAY,
        TEXT=TEXT,
        MUTED=MUTED,
        BG=BG,
        WHITE=WHITE,
        BORDER=BORDER,
        TINT=TINT
    )
